import logging
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

logger = logging.getLogger(__name__)


class ExcelReader:
    """Excel 文件读取器"""
    
    def __init__(self, file_path: str):
        """
        初始化 Excel 读取器
        
        Args:
            file_path: Excel 文件路径
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
    
    def load(self) -> bool:
        """
        加载 Excel 文件
        
        Returns:
            bool: 是否成功加载
        """
        try:
            if not os.path.exists(self.file_path):
                logger.error(f"文件不存在: {self.file_path}")
                return False
            
            self.workbook = load_workbook(self.file_path, read_only=True)
            self.worksheet = self.workbook.active
            logger.info(f"成功加载 Excel 文件: {self.file_path}")
            return True
            
        except Exception as e:
            logger.error(f"加载 Excel 文件失败: {str(e)}")
            return False
    
    def get_column_data(self, column: str = 'A', 
                      start_row: int = 2,
                      end_row: Optional[int] = None) -> List[Any]:
        if not self.worksheet:
            logger.error("工作表未加载")
            return []
        
        try:
            from openpyxl.utils import column_index_from_string
            col_idx = column_index_from_string(column)
            
            data = []
            chunk_size = 5000
            current_row = start_row
            
            logger.info(f"开始读取列 {column}，从行 {start_row}")
            
            while True:
                batch_end = current_row + chunk_size - 1
                rows = self.worksheet.iter_rows(
                    min_row=current_row, max_row=batch_end,
                    min_col=col_idx, max_col=col_idx,
                    values_only=True
                )
                batch = [row[0] for row in rows if row[0] is not None]
                
                if not batch:
                    break
                
                data.extend(batch)
                current_row = batch_end + 1
                
                if len(batch) < chunk_size:
                    break
            
            logger.info(f"从列 {column} 读取到 {len(data)} 条数据")
            return data
            
        except Exception as e:
            logger.error(f"读取列数据失败: {str(e)}")
            import traceback
            logger.error(f"错误堆栈: {traceback.format_exc()}")
            return []
    
    def get_multiple_columns(self, columns: List[str],
                           start_row: int = 2,
                           end_row: Optional[int] = None) -> Dict[str, List[Any]]:
        if not self.worksheet:
            logger.error("工作表未加载")
            return {}
        
        try:
            from openpyxl.utils import column_index_from_string
            
            data_dict = {col: [] for col in columns}
            col_indices = []
            for column in columns:
                column_letter = self.get_column_letter_by_name(column)
                if column_letter is None:
                    column_letter = column
                col_idx = column_index_from_string(column_letter)
                col_indices.append((column, col_idx))
            
            min_col = min(idx for _, idx in col_indices)
            max_col = max(idx for _, idx in col_indices)
            
            chunk_size = 5000
            current_row = start_row
            
            while True:
                batch_end = current_row + chunk_size - 1
                rows = self.worksheet.iter_rows(
                    min_row=current_row, max_row=batch_end,
                    min_col=min_col, max_col=max_col,
                    values_only=True
                )
                
                batch_count = 0
                for row in rows:
                    batch_count += 1
                    for column, col_idx in col_indices:
                        cell_val = row[col_idx - min_col] if col_idx - min_col < len(row) else None
                        if cell_val is not None:
                            data_dict[column].append(cell_val)
                
                if batch_count == 0:
                    break
                current_row = batch_end + 1
                
                if batch_count < chunk_size:
                    break
            
            for column in columns:
                logger.info(f"从列 {column} 读取到 {len(data_dict[column])} 条数据")
            return data_dict
            
        except Exception as e:
            logger.error(f"读取多列数据失败: {str(e)}")
            return {}
    
    def get_row_count(self) -> int:
        """获取数据行数（不包括标题）"""
        if not self.worksheet:
            return 0
        return max(0, self.worksheet.max_row - 1)
    
    def get_column_names(self, row: int = 1) -> List[str]:
        if not self.worksheet:
            return []

        try:
            column_names = []
            for r in self.worksheet.iter_rows(min_row=row, max_row=row, values_only=False):
                for cell in r:
                    if cell.value is not None and str(cell.value).strip():
                        column_names.append(str(cell.value).strip())

            return column_names

        except Exception as e:
            logger.error(f"获取列名失败: {str(e)}")
            return []
    
    def get_column_index_by_name(self, column_name: str) -> Optional[int]:
        if not self.worksheet:
            return None

        try:
            for r in self.worksheet.iter_rows(min_row=1, max_row=1, values_only=False):
                for cell in r:
                    if cell.value and str(cell.value).strip() == column_name:
                        return cell.column
            return None
        except Exception as e:
            logger.error(f"查找列索引失败: {str(e)}")
            return None
    
    def get_column_letter_by_name(self, column_name: str) -> Optional[str]:
        if not self.worksheet:
            return None

        try:
            from openpyxl.utils import get_column_letter
            for r in self.worksheet.iter_rows(min_row=1, max_row=1, values_only=False):
                for cell in r:
                    if cell.value and str(cell.value).strip() == column_name:
                        return get_column_letter(cell.column)
            return None
        except Exception as e:
            logger.error(f"查找列字母失败: {str(e)}")
            return None
    
    def validate_data(self, data: List[Any]) -> Dict[str, Any]:
        validation_result = {
            'total': len(data),
            'valid': 0,
            'invalid': 0,
            'empty': 0,
            'errors': []
        }
        
        max_errors = 100
        for idx, value in enumerate(data, 1):
            if value is None or value == '':
                validation_result['empty'] += 1
                if len(validation_result['errors']) < max_errors:
                    validation_result['errors'].append({
                        'row': idx,
                        'value': value,
                        'error': '空值'
                    })
            elif isinstance(value, (int, float)):
                validation_result['valid'] += 1
            elif isinstance(value, str) and value.strip():
                validation_result['valid'] += 1
            else:
                validation_result['invalid'] += 1
                if len(validation_result['errors']) < max_errors:
                    validation_result['errors'].append({
                        'row': idx,
                        'value': value,
                        'error': '无效数据类型'
                    })
        
        logger.info(f"数据验证完成: 总计 {validation_result['total']}, "
                   f"有效 {validation_result['valid']}, "
                   f"无效 {validation_result['invalid']}, "
                   f"空值 {validation_result['empty']}")
        
        return validation_result
    
    def get_file_info(self) -> Dict[str, Any]:
        if not self.workbook:
            return {}

        try:
            col_names = self.get_column_names()
            max_row = self.worksheet.max_row
            data_rows = max(0, max_row - 1)

            return {
                'file_path': self.file_path,
                'sheet_name': self.worksheet.title,
                'total_rows': max_row,
                'total_columns': len(col_names),
                'column_names': col_names,
                'data_rows': data_rows
            }
        except Exception as e:
            logger.error(f"获取文件信息失败: {str(e)}")
            return {}
    
    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
            logger.info("Excel 文件已关闭")