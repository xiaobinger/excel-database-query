import logging
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Excel 文件写入器"""
    
    def __init__(self, file_path: str):
        """
        初始化 Excel 写入器
        
        Args:
            file_path: Excel 文件路径
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
    
    def load_existing(self) -> bool:
        """
        加载已存在的 Excel 文件
        
        Returns:
            bool: 是否成功加载
        """
        try:
            if os.path.exists(self.file_path):
                self.workbook = load_workbook(self.file_path)
                logger.info(f"成功加载已存在的 Excel 文件: {self.file_path}")
                return True
            else:
                logger.warning(f"文件不存在，将创建新文件: {self.file_path}")
                return False
        except Exception as e:
            logger.error(f"加载 Excel 文件失败: {str(e)}")
            return False
    
    def create_new(self):
        """创建新的 Excel 工作簿"""
        self.workbook = Workbook()
        logger.info(f"创建新的 Excel 工作簿: {self.file_path}")
    
    def create_sheet(self, sheet_name: str) -> bool:
        """
        创建新的工作表
        
        Args:
            sheet_name: 工作表名称
        
        Returns:
            bool: 是否成功创建
        """
        try:
            if not self.workbook:
                self.create_new()
            
            # 检查工作表是否已存在
            if sheet_name in [sheet.title for sheet in self.workbook.worksheets]:
                logger.warning(f"工作表 {sheet_name} 已存在，将删除")
                self.workbook.remove(self.workbook[sheet_name])
            
            self.worksheet = self.workbook.create_sheet(sheet_name)
            logger.info(f"成功创建工作表: {sheet_name}")
            return True
            
        except Exception as e:
            logger.error(f"创建工作表失败: {str(e)}")
            return False
    
    def write_headers(self, headers: List[str], 
                    bold: bool = True,
                    bg_color: str = '4472C4',
                    font_color: str = 'FFFFFF'):
        """
        写入表头
        
        Args:
            headers: 表头列表
            bold: 是否加粗
            bg_color: 背景颜色
            font_color: 字体颜色
        """
        if not self.worksheet:
            logger.error("工作表未创建")
            return False
        
        try:
            header_font = Font(bold=bold, color=font_color, size=11)
            header_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col_idx, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            logger.info(f"成功写入 {len(headers)} 个表头")
            return True
            
        except Exception as e:
            logger.error(f"写入表头失败: {str(e)}")
            return False
    
    def write_data(self, data: List[List[Any]], 
                  start_row: int = 2,
                  format_dict: Optional[Dict[str, Any]] = None):
        """
        写入数据
        
        Args:
            data: 数据列表（二维列表）
            start_row: 开始行号
            format_dict: 格式化字典 {列名: 格式配置}
        """
        if not self.worksheet:
            logger.error("工作表未创建")
            return False
        
        try:
            data_font = Font(size=10)
            data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row_idx, row_data in enumerate(data, start_row):
                for col_idx, cell_value in enumerate(row_data, 1):
                    cell = self.worksheet.cell(row=row_idx, column=col_idx)
                    
                    # 应用格式化
                    if format_dict and col_idx <= len(format_dict):
                        col_name = list(format_dict.keys())[col_idx - 1]
                        cell_value = self._format_value(cell_value, format_dict[col_name])
                    else:
                        cell_value = self._format_value(cell_value)
                    
                    cell.value = cell_value
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
            
            logger.info(f"成功写入 {len(data)} 行数据")
            return True
            
        except Exception as e:
            logger.error(f"写入数据失败: {str(e)}")
            return False
    
    def get_column_index_by_name(self, column_name: str, row_num: int = 1) -> Optional[int]:
        """
        根据列名获取列索引
        
        Args:
            column_name: 列名
            row_num: 标题行号
            
        Returns:
            列索引（从1开始），如果未找到返回 None
        """
        if not self.worksheet:
            logger.error("工作表未创建")
            return None
        
        try:
            for col_idx in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=row_num, column=col_idx)
                if cell.value == column_name:
                    logger.info(f"找到列 {column_name} 在索引 {col_idx}")
                    return col_idx
            logger.warning(f"未找到列 {column_name}")
            return None
        except Exception as e:
            logger.error(f"查找列索引失败: {str(e)}")
            return None
    
    def get_row_by_value(self, column_index: int, value: Any, start_row: int = 2) -> Optional[int]:
        """
        根据列值查找行索引
        
        Args:
            column_index: 列索引
            value: 要查找的值
            start_row: 开始行号
            
        Returns:
            行索引，如果未找到返回 None
        """
        if not self.worksheet:
            logger.error("工作表未创建")
            return None
        
        try:
            for row_idx in range(start_row, self.worksheet.max_row + 1):
                cell = self.worksheet.cell(row=row_idx, column=column_index)
                if cell.value == value:
                    logger.info(f"找到值 {value} 在行 {row_idx}")
                    return row_idx
            logger.warning(f"未找到值 {value}")
            return None
        except Exception as e:
            logger.error(f"查找行索引失败: {str(e)}")
            return None
    
    def write_data_by_column_name(self, results: List[Dict[str, Any]], 
                               param_column_name: str, 
                               header_row: int = 1, 
                               data_start_row: int = 2):
        """
        根据列名和参数值写入数据
        
        Args:
            results: 查询结果列表
            param_column_name: 参数列名
            header_row: 标题行号
            data_start_row: 数据开始行号
            
        Returns:
            是否成功写入
        """
        if not self.worksheet:
            logger.error("工作表未创建")
            return False
        
        try:
            # 获取参数列的索引
            param_col_idx = self.get_column_index_by_name(param_column_name, header_row)
            if not param_col_idx:
                logger.error(f"未找到参数列 {param_column_name}")
                return False
            
            # 获取所有结果列名
            if not results:
                logger.warning("没有结果数据")
                return True
            
            # 构建结果字典 {参数值: 结果字典}，支持多种参数值类型
            result_dict = {}
            for result in results:
                if isinstance(result, dict):
                    # 提取参数值
                    param_value = result.get(param_column_name)
                    if param_value is not None:
                        # 处理不可哈希类型（如列表）
                        if isinstance(param_value, list):
                            # 将列表转换为元组作为键
                            tuple_key = tuple(param_value)
                            result_dict[tuple_key] = result
                            # 同时保存字符串表示作为键
                            str_key = str(param_value)
                            result_dict[str_key] = result
                        else:
                            # 同时保存原始类型和字符串类型的参数值作为键
                            result_dict[param_value] = result
                            if not isinstance(param_value, str):
                                result_dict[str(param_value)] = result
                            # 保存去除空格后的字符串作为键
                            if isinstance(param_value, str):
                                result_dict[param_value.strip()] = result
            
            # 获取所有需要写入的列名
            all_result_columns = set()
            for result in results:
                all_result_columns.update(result.keys())
            
            # 排除参数列
            all_result_columns.discard(param_column_name)
            all_result_columns = sorted(all_result_columns)
            
            # 记录已写入的列索引
            column_mapping = {}
            for col_name in all_result_columns:
                col_idx = self.get_column_index_by_name(col_name, header_row)
                if col_idx:
                    column_mapping[col_name] = col_idx
            
            # 如果没有匹配到任何列，跳过
            if not column_mapping:
                logger.error("没有匹配到任何结果列，跳过写入")
                return False
            
            # 遍历所有行，根据参数值填充数据
            max_row = self.worksheet.max_row
            matched_count = 0
            
            for row_idx in range(data_start_row, max_row + 1):
                # 获取当前行的参数值
                param_cell = self.worksheet.cell(row=row_idx, column=param_col_idx)
                param_value = param_cell.value
                
                if param_value is None:
                    continue
                
                # 尝试多种方式匹配参数值
                result = None
                if param_value in result_dict:
                    # 原始类型匹配
                    result = result_dict[param_value]
                elif str(param_value) in result_dict:
                    # 字符串类型匹配
                    result = result_dict[str(param_value)]
                elif isinstance(param_value, str) and param_value.strip() in result_dict:
                    # 去除空格后匹配
                    result = result_dict[param_value.strip()]
                
                if result:
                    matched_count += 1
                    # 写入结果数据到对应列
                    for col_name, col_idx in column_mapping.items():
                        if col_name in result:
                            cell = self.worksheet.cell(row=row_idx, column=col_idx)
                            cell.value = result[col_name]
            
            logger.info(f"数据写入完成，共匹配到 {matched_count} 行数据")
            return True
        except Exception as e:
            logger.error(f"写入数据失败: {str(e)}")
            return False
    
    def _format_value(self, value: Any, 
                   format_config: Optional[Dict[str, Any]] = None) -> Any:
        """
        格式化单元格值
        
        Args:
            value: 原始值
            format_config: 格式配置
        
        Returns:
            格式化后的值
        """
        if value is None:
            return ''
        
        if not format_config:
            return value
        
        try:
            # 日期格式化
            if format_config.get('type') == 'datetime':
                if isinstance(value, datetime):
                    date_format = format_config.get('format', '%Y-%m-%d %H:%M:%S')
                    return value.strftime(date_format)
            
            # 货币格式化
            elif format_config.get('type') == 'currency':
                if isinstance(value, (int, float)):
                    symbol = format_config.get('symbol', '¥')
                    decimals = format_config.get('decimals', 2)
                    return f"{symbol}{value:.{decimals}f}"
            
            # 百分比格式化
            elif format_config.get('type') == 'percentage':
                if isinstance(value, (int, float)):
                    return f"{value:.2f}%"
            
            # 映射格式化
            elif format_config.get('type') == 'mapping':
                mapping = format_config.get('mapping', {})
                return mapping.get(value, value)
            
            return value
            
        except Exception as e:
            logger.warning(f"格式化值失败: {value}, {str(e)}")
            return value
    
    def write_query_results(self, results: List[Dict[str, Any]], 
                         sheet_name: str = '查询结果',
                         include_metadata: bool = True,
                         new_sheet: bool = True):
        """
        写入查询结果
        
        Args:
            results: 查询结果列表
            sheet_name: 工作表名称
            include_metadata: 是否包含元数据（查询参数、执行时间等）
            new_sheet: 是否新建工作表，False 则填充到当前工作表
        """
        try:
            # 创建或使用当前工作表
            if new_sheet:
                # 创建工作表
                if not self.create_sheet(sheet_name):
                    return False
            else:
                # 使用当前工作表
                if not self.workbook:
                    logger.error("工作簿未加载")
                    return False
                # 使用第一个工作表作为当前工作表
                self.worksheet = self.workbook.active
                logger.info(f"使用当前工作表: {self.worksheet.title}")
            
            if not results:
                logger.warning("没有结果需要写入")
                return True
            
            # 获取列名
            first_result = results[0]
            if include_metadata and 'result' in first_result:
                # 包含元数据的情况
                headers = ['查询参数', '查询状态', '执行时间'] + list(first_result['result'][0].keys()) if first_result['result'] else []
            else:
                # 仅结果数据的情况
                headers = list(first_result.keys()) if 'result' not in first_result else list(first_result['result'][0].keys()) if first_result['result'] else []
            
            # 写入表头
            self.write_headers(headers)
            
            # 写入数据
            data = []
            for result in results:
                if include_metadata and 'result' in result:
                    # 包含元数据
                    params_str = str(result.get('params', {}))
                    status = '成功' if result.get('success') else '失败'
                    exec_time = result.get('execution_time', 'N/A')
                    result_data = list(result['result'][0]) if result['result'] else []
                    row = [params_str, status, exec_time] + result_data
                else:
                    # 仅结果数据
                    row = list(result.values()) if 'result' not in result else list(result['result'][0]) if result['result'] else []
                
                data.append(row)
            
            self.write_data(data)
            return True
            
        except Exception as e:
            logger.error(f"写入查询结果失败: {str(e)}")
            return False
    
    def write_summary(self, summary: Dict[str, Any], 
                   sheet_name: str = '查询统计'):
        """
        写入统计摘要
        
        Args:
            summary: 统计数据字典
            sheet_name: 工作表名称
        """
        try:
            # 创建工作表
            if not self.create_sheet(sheet_name):
                return False
            
            # 写入统计数据
            data = []
            for key, value in summary.items():
                data.append([key, str(value)])
            
            self.write_headers(['统计项', '值'])
            self.write_data(data)
            return True
            
        except Exception as e:
            logger.error(f"写入统计摘要失败: {str(e)}")
            return False
    
    def auto_adjust_column_width(self):
        """自动调整列宽"""
        if not self.worksheet:
            return
        
        try:
            # 只遍历前100行数据来计算列宽，提高性能
            max_rows_to_check = min(100, self.worksheet.max_row)
            
            for column in self.worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                # 只遍历前max_rows_to_check行
                for cell in column[:max_rows_to_check]:
                    try:
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                self.worksheet.column_dimensions[column_letter].width = adjusted_width
            
            logger.info("列宽自动调整完成")
            return True
            
        except Exception as e:
            logger.error(f"调整列宽失败: {str(e)}")
            return False
    
    def freeze_panes(self, row: int = 1, col: int = 1):
        """
        冻结窗格
        
        Args:
            row: 冻结行数
            col: 冻结列数
        """
        if not self.worksheet:
            return
        
        try:
            # 将行列转换为单元格地址（如 A1, B2 等）
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col)
            cell_address = f"{col_letter}{row}"
            
            self.worksheet.freeze_panes = cell_address
            logger.info(f"冻结窗格: {cell_address}")
            return True
        except Exception as e:
            logger.error(f"冻结窗格失败: {str(e)}")
            return False
    
    def save(self) -> bool:
        """
        保存 Excel 文件
        
        Returns:
            bool: 是否成功保存
        """
        try:
            if not self.workbook:
                logger.error("工作簿未初始化")
                return False
            
            # 自动调整列宽
            self.auto_adjust_column_width()
            
            # 冻结首行
            self.freeze_panes()
            
            # 删除默认工作表（如果存在）
            if 'Sheet' in [sheet.title for sheet in self.workbook.worksheets]:
                if len(self.workbook.worksheets) > 1:
                    self.workbook.remove(self.workbook['Sheet'])
            
            # 保存文件
            self.workbook.save(self.file_path)
            logger.info(f"成功保存 Excel 文件: {self.file_path}")
            return True
            
        except Exception as e:
            logger.error(f"保存 Excel 文件失败: {str(e)}")
            return False
    
    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
            logger.info("Excel 工作簿已关闭")
    
    def get_active_sheet(self):
        """
        获取活动工作表（第一个工作表）
        
        Returns:
            工作表对象
        """
        if not self.workbook:
            logger.error("工作簿未初始化")
            return None
        
        # 返回第一个工作表
        if self.workbook.worksheets:
            self.worksheet = self.workbook.worksheets[0]
            logger.info(f"使用工作表: {self.worksheet.title}")
            return self.worksheet
        else:
            logger.error("工作簿中没有工作表")
            return None
    
    def append_columns_to_sheet(self, headers: List[str], data: List[List[Any]]) -> bool:
        """
        在现有工作表中追加列
        
        Args:
            headers: 要添加的列标题列表
            data: 要添加的数据列表（二维列表）
        
        Returns:
            是否成功
        """
        try:
            if not self.worksheet:
                logger.error("工作表未设置")
                return False
            
            # 获取当前工作表的最后一行
            max_row = self.worksheet.max_row
            
            # 获取当前工作表的最后一列
            max_col = self.worksheet.max_column
            
            # 写入新的列标题（在第一行）
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col_idx, header in enumerate(headers, 1):
                col_letter = chr(64 + max_col + col_idx)
                cell = self.worksheet.cell(row=1, column=max_col + col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 写入数据行
            data_font = Font(size=10)
            data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, cell_value in enumerate(row_data, 1):
                    col_letter = chr(64 + max_col + col_idx)
                    cell = self.worksheet.cell(row=row_idx, column=max_col + col_idx)
                    cell.value = cell_value
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
            
            logger.info(f"成功在工作表 {self.worksheet.title} 中添加 {len(headers)} 列，{len(data)} 行数据")
            return True
            
        except Exception as e:
            logger.error(f"追加列失败: {str(e)}")
            return False