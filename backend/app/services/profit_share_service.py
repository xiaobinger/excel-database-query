"""代理商分润计算与导出服务

根据两个SQL脚本查询数据：
  脚本1: 查询代理关系及代理相关成本费率
  脚本2: 查询交易订单（含直属代理编号、一级代理商编号）

计算逻辑（从直属代理往上逐级计算）：
  1. 根据订单的 终端类型、产品类型 找到代理配置组
  2. 通过订单的直属代理编号（为空则取一级代理商编号）匹配成本配置
  3. 总分润池 = 交易金额*(交易费率-一级代理费率成本) + (交易T0费-一级代理T0成本)
  4. 直属代理分润 = 交易金额*(交易费率-直属代理费率成本) + (交易T0费-直属代理T0成本)
  5. 往上逐级分润 = 交易金额*(下级费率成本-上级费率成本) + (下级T0成本-上级T0成本)，分配给上级
  6. 所有代理分润总和不超过总分润池
  7. 特殊规则: 交易金额 > 1000 时，费率成本取刷卡贷记卡相关值
  8. 借记卡封顶: 当 交易金额*交易费率 > 封顶值(分/100=元) 时，交易金额替换为 交易手续费/交易费率
"""

import json
import logging
import os
import threading
import uuid
from collections import defaultdict
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from app import db
from app.models.database import DatabaseConnection
from app.models.query_task import QueryTask
from app.utils.helpers import beijing_isoformat

logger = logging.getLogger(__name__)

_task_progress = {}
_task_lock = threading.Lock()
_task_threads = {}
_task_cancel_events = {}

# ── SQL 脚本 ──────────────────────────────────────────────

SQL_AGENT_RATE_CONFIG = """
SELECT
    ac.agent_no       AS 代理商编号,
    IFNULL(IFNULL(a.agent_name, ai.agent_value_code), a.login_phone) AS 代理商名称,
    a.rank            AS 代理商等级,
    a.login_phone     AS 代理手机号码,
    a.parent_agent_no AS 上级代理商编号,
    p.agent_name      AS 所属上级,
    r.agent_name      AS 所属一级,
    ac.device_type    AS 终端类型,
    ac.product_type   AS 产品类型,
    CASE WHEN ac.rate_type = 0 THEN '代理费率成本' WHEN ac.rate_type = 2 THEN '代理T0成本' END AS 费率类型,
    ac.product_id     AS 产品ID,
    ac.rate_info      AS 成本内容
FROM pro_isp_db.tbl_agent_rate_config ac
LEFT JOIN pro_isp_db.tbl_agent a  ON ac.agent_no = a.agent_no
LEFT JOIN pro_isp_db.tbl_agent p  ON a.parent_agent_no = p.agent_no
LEFT JOIN pro_isp_db.tbl_agent r  ON a.root_agent_no = r.agent_no
LEFT JOIN posp_business.org_migrate_mapping om ON om.old_org_code = r.agent_no
LEFT JOIN pro_isp_db.tbl_agent_item ai ON ai.agent_no = a.agent_no AND ai.agent_key_code = 'legalPersonName'
WHERE ac.product_type IN (1, 4, 8, 9)
  AND ac.rate_type != 1
  AND ac.`status` = 0
  AND ac.product_id IS NOT NULL
  AND r.agent_no = :org_no
"""

SQL_TRADE_ORDERS = """
SELECT
    o.order_no                                AS 订单号,
    o.trade_time                              AS 交易时间,
    IFNULL(td.agent_no, om.old_org_code)      AS 所属代理编号,
    om.old_org_code                           AS 一级代理商编号,
    o.trade_amount / 100                      AS 交易金额,
    o.trade_rate / 100                        AS 交易费率,
    o.trade_fee_amount / 100                  AS 交易手续费,
    IFNULL(o.trade_t0_fee, 0) / 100           AS 交易T0服务费,
    IF(o.channel_rate > 0.1, o.channel_rate / 100, o.channel_rate) AS 服务商费率成本,
    IFNULL(o.channel_t0_fee, 0) / 100         AS 服务商T0成本,
    IF(o.org_rate > 0.1, o.org_rate / 100, o.org_rate) AS 一级代理费率成本,
    IFNULL(o.org_t0_fee, 0) / 100             AS 一级代理T0成本,
    CASE o.trade_type
        WHEN 1 THEN '刷卡'
        WHEN 2 THEN '银二'
        WHEN 3 THEN '手机PAY'
        WHEN 4 THEN '支付宝'
        WHEN 5 THEN '微信'
    END                                       AS 交易类型,
    IF(o.card_type = 1, '借记卡', '贷记卡')     AS 卡类型,
    opm.old_product_id                        AS 产品ID,
    CASE WHEN opm.old_product_type = 1 THEN 1
         WHEN opm.old_product_type = 3 THEN 8
         WHEN opm.old_product_type = 4 THEN 9
         WHEN opm.old_product_type = 2 THEN 4
    END                                       AS 产品类型,
    d.device_type                             AS 终端类型
FROM posp_business.org_migrate_mapping om
INNER JOIN posp_business.trade_order o
    ON o.org_no = om.new_org_code
    AND o.trade_status = 1
    AND o.trade_time BETWEEN :start_time AND :end_time
LEFT JOIN posp_business.device d
    ON o.device_sn = d.device_sn
LEFT JOIN posp_business.org_product_migrate_mapping opm
    ON d.product_code = opm.new_product_id
LEFT JOIN pro_mcht_db.tbl_device td
    ON td.device_sn = d.device_sn
    AND td.root_agent_no = :org_no
WHERE om.old_org_code = :org_no
"""

# ── 费率键映射 ────────────────────────────────────────────

# 交易类型/卡类型 → (费率成本键, T0成本键)
# 手机PAY (NFC支付) 按刷卡处理，区分借贷记卡
RATE_KEY_MAP = {
    ('微信', None):     ('wxPayRate',      'tsWxPay'),
    ('支付宝', None):   ('aliPayRate',     'tsAliPay'),
    ('银二', None):     ('unionPayRate',   'tsUnionPay'),
    ('刷卡', '借记卡'): ('debitPayRate',   'tsDebitCardPay'),
    ('刷卡', '贷记卡'): ('creditPayRate',  'tsCreditCardPay'),
    ('手机PAY', '借记卡'): ('debitPayRate', 'tsDebitCardPay'),
    ('手机PAY', '贷记卡'): ('creditPayRate', 'tsCreditCardPay'),
}

# 交易金额 > 1000 时强制使用刷卡贷记卡费率
HIGH_AMOUNT_RATE_KEYS = ('creditPayRate', 'tsCreditCardPay')

# 导出表头
EXPORT_HEADERS = [
    '代理商编号', '代理商名称', '分润金额', '月份',
    '代理商电话', '代理商等级', '所属一级代理商名称',
    '上级代理商', '上级关系'
]

# ── 进度管理 ──────────────────────────────────────────────

def _update_progress(task_id: str, progress: int, log_message: str = None, level: str = 'info'):
    with _task_lock:
        _task_progress[task_id] = {
            'progress': progress,
            'log_message': log_message,
            'level': level,
            'timestamp': beijing_isoformat(datetime.utcnow()),
        }


def _get_progress(task_id: str) -> dict:
    with _task_lock:
        return _task_progress.get(task_id, {'progress': 0})


# ── 数据解析工具 ──────────────────────────────────────────

def _parse_rate_info(rate_info_str) -> dict:
    """解析成本内容 JSON 字符串，返回费率字典"""
    if not rate_info_str:
        return {}
    if isinstance(rate_info_str, dict):
        return rate_info_str
    try:
        return json.loads(rate_info_str)
    except (json.JSONDecodeError, TypeError):
        logger.warning(f"解析 rate_info 失败: {rate_info_str}")
        return {}


def _to_float(value, default=0.0) -> float:
    """安全转换为 float"""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def _get_rate_keys(trade_type: str, card_type: str, trade_amount: float) -> Tuple[str, str]:
    """根据交易类型、卡类型、交易金额确定费率键

    特殊规则: 交易金额 > 1000 时，取刷卡贷记卡相关值
    """
    if trade_amount is not None and trade_amount > 1000:
        return HIGH_AMOUNT_RATE_KEYS

    # 刷卡和手机PAY 需要区分借贷记卡
    if trade_type in ('刷卡', '手机PAY'):
        key = (trade_type, card_type)
    else:
        key = (trade_type, None)
    return RATE_KEY_MAP.get(key, HIGH_AMOUNT_RATE_KEYS)


def _extract_rate(rate_info_dict: dict, rate_key: str) -> float:
    """从费率字典中提取指定键的值"""
    if not rate_info_dict:
        return 0.0
    return _to_float(rate_info_dict.get(rate_key))


# ── 代理配置构建 ──────────────────────────────────────────

class AgentNode:
    """代理节点"""
    __slots__ = ('agent_no', 'agent_name', 'rank', 'login_phone',
                 'parent_agent_no', 'parent_name', 'root_name',
                 'rate_cost_info', 't0_cost_info')

    def __init__(self, agent_no, agent_name, rank, login_phone,
                 parent_agent_no, parent_name, root_name):
        self.agent_no = agent_no
        self.agent_name = agent_name or ''
        self.rank = rank
        self.login_phone = login_phone or ''
        self.parent_agent_no = parent_agent_no
        self.parent_name = parent_name or ''
        self.root_name = root_name or ''
        self.rate_cost_info = {}   # 费率成本 JSON (rate_type=0)
        self.t0_cost_info = {}     # T0成本 JSON (rate_type=2)


def _build_agent_configs(rows: List[dict], org_no: str) -> Dict[Tuple, Dict[str, AgentNode]]:
    """构建代理配置字典

    返回: {(device_type, product_type): {agent_no: AgentNode}}
    """
    configs: Dict[Tuple, Dict[str, AgentNode]] = defaultdict(dict)

    for row in rows:
        agent_no = str(row.get('代理商编号', '')).strip()
        if not agent_no:
            continue

        device_type = str(row.get('终端类型', '')).strip()
        product_type = row.get('产品类型')
        # 产品类型统一转为字符串作为字典键
        product_type_str = str(product_type).strip() if product_type is not None else ''
        rate_type = row.get('费率类型')
        rate_info_str = row.get('成本内容', '')

        key = (device_type, product_type_str)
        node = configs[key].get(agent_no)

        if node is None:
            node = AgentNode(
                agent_no=agent_no,
                agent_name=row.get('代理商名称', ''),
                rank=row.get('代理商等级'),
                login_phone=row.get('代理手机号码', ''),
                parent_agent_no=str(row.get('上级代理商编号', '')).strip() if row.get('上级代理商编号') else None,
                parent_name=row.get('所属上级', ''),
                root_name=row.get('所属一级', ''),
            )
            configs[key][agent_no] = node

        # 合并费率/T0成本信息（费率类型可能是数字 0/2 或字符串 '代理费率成本'/'代理T0成本'）
        parsed = _parse_rate_info(rate_info_str)
        if rate_type in (0, '代理费率成本'):
            node.rate_cost_info = parsed
        elif rate_type in (2, '代理T0成本'):
            node.t0_cost_info = parsed

    return dict(configs)


def _get_ancestor_chain(node: AgentNode, agents: Dict[str, AgentNode], order: dict = None) -> List[str]:
    """获取节点的上级名称链（从根到当前代理，含当前代理自己）

    处理一级代理不在 agents 中的情况：用 order 中的"所属一级"名称补全。
    """
    chain = []
    current = node
    visited = set()

    org_agent_no = str(order.get('一级代理商编号', '')).strip() if order and order.get('一级代理商编号') else ''

    while current and current.parent_agent_no and current.parent_agent_no not in visited:
        visited.add(current.parent_agent_no)
        parent_no = current.parent_agent_no
        parent = agents.get(parent_no)
        if parent:
            chain.append(parent.agent_name or parent_no)
            current = parent
        elif order and parent_no == org_agent_no:
            # 一级代理不在 agents 中，用订单中的"所属一级"名称
            chain.append(order.get('所属一级', '') or parent_no)
            break
        else:
            break

    chain.reverse()
    # 末尾加上当前代理自己
    chain.append(node.agent_name or node.agent_no)
    return chain


# ── 分润计算 ──────────────────────────────────────────────

def _calculate_order_shares(
    order: dict,
    agents: Dict[str, AgentNode],
) -> Tuple[Dict[str, float], str, float]:
    """计算单笔订单各代理的分润金额（从直属代理往上逐级计算）

    1. 总分润池 = 交易金额*(交易费率 - 一级代理费率成本) + (交易T0费 - 一级代理T0成本)
    2. 直属代理分润 = 交易金额*(交易费率 - 直属代理费率成本) + (交易T0费 - 直属代理T0成本)
    3. 往上逐级分润 = 交易金额*(下级代理费率成本 - 上级代理费率成本) + (下级T0成本 - 上级代理T0成本)
    4. 分润分配给上级代理，累计不超过总分润池
    5. 一级代理费率成本/T0成本 优先用订单自带的值（最准确）
    6. 借记卡封顶：当交易金额*交易费率 > 封顶值(分/100=元)时，交易金额替换为 封顶值(元)/(交易费率-一级代理费率成本)

    返回: (shares: {agent_no: share_amount}, skip_reason: str, total_pool: float)
        skip_reason 为空表示正常计算有结果；非空则表示跳过原因
        total_pool 为该笔订单的总分润池（用于订单明细展示）
    """
    trade_amount = _to_float(order.get('交易金额'))
    trade_rate = _to_float(order.get('交易费率'))
    trade_fee_amount = _to_float(order.get('交易手续费'))
    trade_t0_fee = _to_float(order.get('交易T0服务费'))
    trade_type = order.get('交易类型', '')
    card_type = order.get('卡类型', '')

    # 一级代理商编号及费率成本（订单自带）
    org_agent_no = str(order.get('一级代理商编号', '')).strip() if order.get('一级代理商编号') else ''
    org_rate = _to_float(order.get('一级代理费率成本'))
    org_t0 = _to_float(order.get('一级代理T0成本'))

    # 直属代理编号，为空则取一级代理商编号
    direct_agent_no = str(order.get('所属代理编号', '')).strip() if order.get('所属代理编号') else ''
    if not direct_agent_no:
        direct_agent_no = org_agent_no

    if not direct_agent_no:
        return {}, '无直属代理编号且无一级代理商编号', 0.0

    rate_key, t0_key = _get_rate_keys(trade_type, card_type, trade_amount)

    # 找到直属代理节点
    direct_node = agents.get(direct_agent_no)
    if not direct_node:
        return {}, f'直属代理未找到配置 (直属代理编号={direct_agent_no})', 0.0

    direct_rate = _extract_rate(direct_node.rate_cost_info, rate_key)
    direct_t0 = _extract_rate(direct_node.t0_cost_info, t0_key)

    # 借记卡封顶处理：当为借记卡交易且触发封顶时，替换交易金额
    # 判断：交易金额*交易费率(元) > 封顶值(分/100=元)
    # 触发后：effective_amount = 交易手续费 / 交易费率
    effective_amount = trade_amount
    if card_type == '借记卡':
        debit_pay_max = _to_float(direct_node.rate_cost_info.get('debitPayMax'))
        if debit_pay_max > 0:
            debit_pay_max_yuan = debit_pay_max / 100
            if trade_amount * trade_rate > debit_pay_max_yuan:
                if trade_rate > 0:
                    effective_amount = trade_fee_amount / trade_rate

    # 总分润池 = 交易金额*(交易费率 - 一级代理费率成本) + (交易T0费 - 一级代理T0成本)
    total_pool = effective_amount * (trade_rate - org_rate) + (trade_t0_fee - org_t0)

    if total_pool <= 0:
        return {}, (f'总分润池<=0 (pool={total_pool:.4f}, 交易金额={trade_amount:.2f}, '
                    f'交易费率={trade_rate}, 一级代理费率成本={org_rate}, '
                    f'交易T0费={trade_t0_fee:.2f}, 一级代理T0成本={org_t0})'), 0.0

    shares: Dict[str, float] = {}
    cumulative = 0.0

    # 直属代理分润 = 交易金额*(交易费率 - 直属代理费率成本) + (交易T0费 - 直属代理T0成本)
    direct_share = effective_amount * (trade_rate - direct_rate) + (trade_t0_fee - direct_t0)

    if direct_share > 0:
        if cumulative + direct_share > total_pool:
            remaining = total_pool - cumulative
            if remaining > 0:
                shares[direct_agent_no] = remaining
                cumulative = total_pool
        else:
            shares[direct_agent_no] = direct_share
            cumulative += direct_share

    # 逐级往上：上级分润 = 交易金额*(下级费率成本 - 上级费率成本) + (下级T0成本 - 上级T0成本)
    # 分润分配给上级代理
    current_rate = direct_rate
    current_t0 = direct_t0
    current_node = direct_node
    visited = {direct_agent_no}

    while cumulative < total_pool:
        parent_no = current_node.parent_agent_no
        if not parent_no or parent_no in visited:
            break

        # 一级代理：优先用订单中的费率成本（最准确），不论是否在 agents 中
        if parent_no == org_agent_no:
            parent_rate = org_rate
            parent_t0 = org_t0
            parent_node = None  # 标记为一级代理，不需要继续往上
        else:
            parent_node = agents.get(parent_no)
            if not parent_node:
                break
            parent_rate = _extract_rate(parent_node.rate_cost_info, rate_key)
            parent_t0 = _extract_rate(parent_node.t0_cost_info, t0_key)

        visited.add(parent_no)

        # 上级分润 = 交易金额*(下级费率成本 - 上级费率成本) + (下级T0成本 - 上级T0成本)
        share = effective_amount * (current_rate - parent_rate) + (current_t0 - parent_t0)

        if share > 0:
            if cumulative + share > total_pool:
                remaining = total_pool - cumulative
                if remaining > 0:
                    shares[parent_no] = remaining
                    cumulative = total_pool
                break
            else:
                shares[parent_no] = share
                cumulative += share

        # 一级代理已到顶级，结束
        if parent_node is None:
            break

        # 继续往上
        current_node = parent_node
        current_rate = parent_rate
        current_t0 = parent_t0

    # 兜底：如果一级代理还没被计算分润，且还有剩余分润池，直接计算一级代理分润
    if cumulative < total_pool and org_agent_no and org_agent_no not in shares and org_agent_no not in visited:
        share = effective_amount * (current_rate - org_rate) + (current_t0 - org_t0)
        if share > 0:
            if cumulative + share > total_pool:
                remaining = total_pool - cumulative
                if remaining > 0:
                    shares[org_agent_no] = remaining
                    cumulative = total_pool
            else:
                shares[org_agent_no] = share
                cumulative += share

    if not shares:
        return {}, (f'各级分润均<=0 (直属代理={direct_agent_no}, 直属费率={direct_rate}, '
                    f'直属T0={direct_t0}, 总分润池={total_pool:.4f})'), 0.0

    return shares, '', total_pool


# ── Excel 生成 ───────────────────────────────────────────

def _generate_excel(aggregated_data: List[dict], output_path: str, order_details: List[dict] = None):
    """生成分润 Excel 文件

    Args:
        aggregated_data: 代理分润聚合数据
        output_path: 输出文件路径
        order_details: 订单明细数据（含总分润池列），用于第二张工作表
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '代理分润明细'

    # 表头样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    data_font = Font(size=10)
    data_align = Alignment(horizontal='left', vertical='center')

    # ── 工作表1: 代理分润明细 ──
    for col_idx, header in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(aggregated_data, 2):
        for col_idx, header in enumerate(EXPORT_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = row_data.get(header, '')
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

            if header == '分润金额' and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00'

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column[:min(100, len(aggregated_data) + 1)]:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    ws.freeze_panes = 'A2'

    # ── 工作表2: 订单明细（含总分润池） ──
    if order_details:
        ws2 = wb.create_sheet(title='订单明细')
        order_headers = list(order_details[0].keys())

        for col_idx, header in enumerate(order_headers, 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 金额/费率类列保留小数
        decimal_headers = {'交易金额', '交易费率', '交易手续费', '交易T0服务费',
                           '服务商费率成本', '服务商T0成本', '一级代理费率成本',
                           '一级代理T0成本', '总分润池'}

        for row_idx, row_data in enumerate(order_details, 2):
            for col_idx, header in enumerate(order_headers, 1):
                cell = ws2.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(header, '')
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border

                if header in decimal_headers and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0000' if header in ('交易费率', '服务商费率成本', '一级代理费率成本') else '0.00'

        # 自动调整列宽
        for column in ws2.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column[:min(100, len(order_details) + 1)]:
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    pass
            ws2.column_dimensions[column_letter].width = min(max_length + 2, 50)

        ws2.freeze_panes = 'A2'

    wb.save(output_path)


# ── 任务管理 ──────────────────────────────────────────────

class ProfitShareService:

    @staticmethod
    def find_database_connection(database_connection_id: int = None) -> Optional[DatabaseConnection]:
        """查找数据库连接，优先使用指定ID，否则按名称查找"""
        if database_connection_id:
            conn = DatabaseConnection.query.get(database_connection_id)
            if conn and conn.is_active:
                return conn
        # 按名称查找"融聚商户通(海科)"
        for name_pattern in ['融聚商户通(海科)', '融聚商户通（海科）', '融聚商户通', '海科']:
            conn = DatabaseConnection.query.filter(
                DatabaseConnection.name.like(f'%{name_pattern}%'),
                DatabaseConnection.is_active == True
            ).first()
            if conn:
                return conn
        return None

    @staticmethod
    def create_task(org_no: str, start_time: str, end_time: str,
                    database_connection_id: int = None, created_by: int = None) -> QueryTask:
        task_id = str(uuid.uuid4())
        task = QueryTask(
            task_id=task_id,
            status='pending',
            type='export',
            output_format='sheets',
        )
        task.set_script_ids_json([])
        task.set_params_values({
            'org_no': org_no,
            'start_time': start_time,
            'end_time': end_time,
            'database_connection_id': database_connection_id,
        })

        if database_connection_id:
            task.set_database_ids([database_connection_id])
            task.database_connection_id = database_connection_id

        if created_by:
            task.created_by = created_by

        db.session.add(task)
        db.session.commit()
        return task

    @staticmethod
    def execute_async(task_id: str, org_no: str, start_time: str, end_time: str,
                      database_connection_id: int, output_dir: str, on_complete=None):
        cancel_event = threading.Event()
        _task_cancel_events[task_id] = cancel_event
        thread = threading.Thread(
            target=ProfitShareService._execute_background,
            args=(task_id, org_no, start_time, end_time, database_connection_id, output_dir, on_complete, cancel_event),
            daemon=True
        )
        _task_threads[task_id] = thread
        thread.start()

    @staticmethod
    def _execute_background(task_id: str, org_no: str, start_time: str, end_time: str,
                            database_connection_id: int, output_dir: str,
                            on_complete=None, cancel_event: threading.Event = None):
        try:
            from app import create_app
            app = create_app()
        except Exception:
            from flask import current_app
            app = current_app._get_current_object()

        with app.app_context():
            task = QueryTask.query.filter_by(task_id=task_id).first()
            if not task:
                return

            try:
                task.status = 'running'
                task.started_at = datetime.utcnow()
                task.add_log('开始执行代理分润导出任务')
                task.progress = 5
                db.session.commit()
                _update_progress(task_id, 5, '开始执行代理分润导出任务')

                # 1. 查找数据库连接
                conn = ProfitShareService.find_database_connection(database_connection_id)
                if not conn:
                    raise RuntimeError('未找到融聚商户通(海科)数据库连接，请在系统中配置该数据库连接')
                task.add_log(f'数据库连接: {conn.name}')
                task.set_database_ids([conn.id])
                task.progress = 10
                db.session.commit()
                _update_progress(task_id, 10, f'数据库连接: {conn.name}')

                if cancel_event and cancel_event.is_set():
                    raise RuntimeError('任务已被手动终止')

                # 2. 获取数据库连接器
                from app.utils.connection_pool import ConnectionPoolManager
                pool = ConnectionPoolManager.get_instance()
                connector = pool.get_connector_with_health_check(conn.id)
                if not connector:
                    raise RuntimeError(f'数据库连接失败: {conn.name}')

                # 3. 执行脚本1: 查询代理费率配置
                task.add_log('查询代理费率配置...')
                task.progress = 15
                db.session.commit()
                _update_progress(task_id, 15, '查询代理费率配置...')

                agent_rows = connector.execute_query(
                    SQL_AGENT_RATE_CONFIG,
                    {'org_no': org_no},
                    timeout=120,
                    chunk_size=5000
                )

                # 将行元组转为字典
                agent_columns = [
                    '代理商编号', '代理商名称', '代理商等级', '代理手机号码',
                    '上级代理商编号', '所属上级', '所属一级', '终端类型',
                    '产品类型', '费率类型', '产品ID', '成本内容'
                ]
                agent_dicts = []
                for row in agent_rows:
                    row_dict = {}
                    for i, col in enumerate(agent_columns):
                        row_dict[col] = row[i] if i < len(row) else None
                    agent_dicts.append(row_dict)

                task.add_log(f'代理费率配置查询完成: {len(agent_dicts)} 行')
                task.progress = 30
                db.session.commit()
                _update_progress(task_id, 30, f'代理费率配置查询完成: {len(agent_dicts)} 行')

                if cancel_event and cancel_event.is_set():
                    raise RuntimeError('任务已被手动终止')

                # 4. 执行脚本2: 查询交易订单
                task.add_log('查询交易订单...')
                task.progress = 35
                db.session.commit()
                _update_progress(task_id, 35, '查询交易订单...')

                order_rows = connector.execute_query(
                    SQL_TRADE_ORDERS,
                    {'org_no': org_no, 'start_time': start_time, 'end_time': end_time},
                    timeout=300,
                    chunk_size=5000
                )

                order_columns = [
                    '订单号', '交易时间', '所属代理编号', '一级代理商编号', '交易金额', '交易费率', '交易手续费', '交易T0服务费',
                    '服务商费率成本', '服务商T0成本', '一级代理费率成本', '一级代理T0成本',
                    '交易类型', '卡类型', '产品ID', '产品类型', '终端类型'
                ]
                order_dicts = []
                for row in order_rows:
                    row_dict = {}
                    for i, col in enumerate(order_columns):
                        row_dict[col] = row[i] if i < len(row) else None
                    order_dicts.append(row_dict)

                task.add_log(f'交易订单查询完成: {len(order_dicts)} 行')
                task.progress = 50
                db.session.commit()
                _update_progress(task_id, 50, f'交易订单查询完成: {len(order_dicts)} 行')

                if not order_dicts:
                    task.add_log('未查询到交易订单数据', 'warning')

                if cancel_event and cancel_event.is_set():
                    raise RuntimeError('任务已被手动终止')

                # 5. 构建代理配置
                task.add_log('构建代理配置...')
                task.progress = 55
                db.session.commit()

                agent_configs = _build_agent_configs(agent_dicts, org_no)
                task.add_log(f'代理配置分组: {len(agent_configs)} 组 (设备类型+产品类型)')

                # 6. 逐笔订单计算分润
                task.add_log('计算分润...')
                task.progress = 60
                db.session.commit()
                _update_progress(task_id, 60, '计算分润...')

                # 聚合: {(agent_no, month): share_sum}
                monthly_shares: Dict[Tuple[str, str], float] = defaultdict(float)
                # 记录代理信息: {agent_no: AgentNode}
                agent_info_map: Dict[str, AgentNode] = {}
                # 记录每个代理的上级关系
                agent_ancestor_chain: Dict[str, str] = {}

                total_orders = len(order_dicts)
                processed = 0
                skipped_no_config = 0
                skipped_no_share = 0
                # 按月份统计: {month: {'total': n, 'skipped': n, 'shared': n}}
                month_stats: Dict[str, Dict[str, int]] = defaultdict(lambda: {'total': 0, 'skipped': 0, 'shared': 0})
                # 记录匹配不到配置的 (代理商编号, 终端类型, 产品类型) 组合，去重
                missed_config_keys = set()
                # 记录无分润结果的跳过原因统计: {reason: count}
                no_share_reasons: Dict[str, int] = defaultdict(int)
                # 记录无分润结果的订单明细（最多保留 20 条用于日志展示）
                no_share_samples: List[str] = []
                NO_SHARE_SAMPLE_LIMIT = 20
                # 收集所有订单明细用于第二张工作表（含总分润池列）
                order_detail_rows: List[dict] = []
                # 按订单号去重，避免重复计算分润
                seen_order_nos: set = set()
                duplicate_count = 0

                for order in order_dicts:
                    # 订单号去重：同一订单号只处理一次
                    order_no = str(order.get('订单号', '')).strip() if order.get('订单号') else ''
                    if order_no:
                        if order_no in seen_order_nos:
                            duplicate_count += 1
                            continue
                        seen_order_nos.add(order_no)

                    device_type = str(order.get('终端类型', '')).strip()
                    product_type = order.get('产品类型')
                    product_type_str = str(product_type).strip() if product_type is not None else ''
                    config_key = (device_type, product_type_str)
                    direct_agent_no = str(order.get('所属代理编号', '')).strip() if order.get('所属代理编号') else ''

                    # 提取月份
                    trade_time = order.get('交易时间')
                    month_str = ''
                    if trade_time:
                        if isinstance(trade_time, datetime):
                            month_str = trade_time.strftime('%Y-%m')
                        elif isinstance(trade_time, str):
                            month_str = trade_time[:7]  # YYYY-MM
                        else:
                            month_str = str(trade_time)[:7]
                    month_stats[month_str]['total'] += 1

                    # 收集订单明细（用于第二张工作表，总分润池先置0，计算分润后更新）
                    detail_row = {k: order.get(k) for k in order_columns}
                    detail_row['总分润池'] = 0
                    order_detail_rows.append(detail_row)

                    agents = agent_configs.get(config_key)
                    if not agents:
                        skipped_no_config += 1
                        month_stats[month_str]['skipped'] += 1
                        missed_config_keys.add((direct_agent_no, device_type, product_type_str))
                        processed += 1
                        continue

                    # 计算分润（从直属代理往上逐级计算）
                    shares, skip_reason, order_total_pool = _calculate_order_shares(order, agents)

                    # 更新订单明细的总分润池（考虑借记卡封顶后的实际值）
                    if order_total_pool > 0:
                        detail_row['总分润池'] = round(order_total_pool, 4)

                    if not shares:
                        skipped_no_share += 1
                        month_stats[month_str]['skipped'] += 1
                        # 直属代理在该配置组中找不到节点
                        if direct_agent_no and direct_agent_no not in agents:
                            missed_config_keys.add((direct_agent_no, device_type, product_type_str))
                        # 收集跳过原因
                        if skip_reason:
                            no_share_reasons[skip_reason] += 1
                            if len(no_share_samples) < NO_SHARE_SAMPLE_LIMIT:
                                trade_time_str = ''
                                if trade_time:
                                    if isinstance(trade_time, datetime):
                                        trade_time_str = trade_time.strftime('%Y-%m-%d %H:%M:%S')
                                    else:
                                        trade_time_str = str(trade_time)
                                trade_amt = _to_float(order.get('交易金额'))
                                sample = (f'  交易时间={trade_time_str} | 直属代理={direct_agent_no} | '
                                          f'一级代理={order.get("一级代理商编号", "")} | '
                                          f'终端类型={device_type} | 产品类型={product_type_str} | '
                                          f'交易类型={order.get("交易类型", "")} | 卡类型={order.get("卡类型", "")} | '
                                          f'交易金额={trade_amt:.2f} | 原因: {skip_reason}')
                                no_share_samples.append(sample)
                        processed += 1
                        continue

                    # 聚合
                    has_positive_share = False
                    for agent_no, share in shares.items():
                        if share == 0:
                            continue
                        monthly_shares[(agent_no, month_str)] += share
                        has_positive_share = True

                        # 记录代理信息
                        if agent_no not in agent_info_map:
                            node = agents.get(agent_no)
                            if node:
                                agent_info_map[agent_no] = node
                                chain = _get_ancestor_chain(node, agents, order)
                                agent_ancestor_chain[agent_no] = '-'.join(chain)
                            else:
                                # 一级代理不在 agents 中，创建虚拟节点
                                root_name = order.get('所属一级', '') or agent_no
                                virtual_node = AgentNode(
                                    agent_no=agent_no,
                                    agent_name=root_name,
                                    rank=1,
                                    login_phone='',
                                    parent_agent_no=None,
                                    parent_name='',
                                    root_name=root_name,
                                )
                                agent_info_map[agent_no] = virtual_node
                                agent_ancestor_chain[agent_no] = root_name

                    if has_positive_share:
                        month_stats[month_str]['shared'] += 1
                    else:
                        skipped_no_share += 1
                        month_stats[month_str]['skipped'] += 1

                    processed += 1
                    if processed % 1000 == 0:
                        task.add_log(f'已处理 {processed}/{total_orders} 笔订单')
                        task.progress = min(60 + int(30 * processed / max(total_orders, 1)), 85)
                        db.session.commit()

                # 输出详细统计
                task.add_log(f'分润计算完成: 查询 {total_orders} 笔, 去重 {duplicate_count} 笔, 实际处理 {processed} 笔订单')
                task.add_log(f'  - 无匹配代理配置: {skipped_no_config} 笔')
                task.add_log(f'  - 无分润结果(总分润池<=0或直属代理未找到): {skipped_no_share} 笔')
                for month_key in sorted(month_stats.keys()):
                    s = month_stats[month_key]
                    task.add_log(f'  月份 {month_key}: 总计 {s["total"]} 笔, 跳过 {s["skipped"]} 笔, 有分润 {s["shared"]} 笔')

                # 输出匹配不到配置的 代理商编号-终端类型-产品类型 清单
                if missed_config_keys:
                    task.add_log(f'  匹配不到配置的组合({len(missed_config_keys)}个, 格式: 代理商编号 | 终端类型 | 产品类型):', 'warning')
                    logger.warning(f'任务 {task_id} 匹配不到配置的组合({len(missed_config_keys)}个):')
                    for agent_no, dev_type, prod_type in sorted(missed_config_keys):
                        line = f'    {agent_no} | {dev_type} | {prod_type}'
                        task.add_log(line, 'warning')
                        logger.warning(line)

                # 输出无分润结果的跳过原因统计 + 样本明细
                if no_share_reasons:
                    task.add_log(f'  无分润结果原因统计({skipped_no_share}笔):', 'warning')
                    logger.warning(f'任务 {task_id} 无分润结果原因统计({skipped_no_share}笔):')
                    # 按出现次数降序
                    for reason, cnt in sorted(no_share_reasons.items(), key=lambda x: -x[1]):
                        line = f'    [{cnt}笔] {reason}'
                        task.add_log(line, 'warning')
                        logger.warning(line)
                    # 输出样本订单明细（最多20条）
                    if no_share_samples:
                        task.add_log(f'  无分润结果订单样本(前{len(no_share_samples)}条):', 'warning')
                        logger.warning(f'任务 {task_id} 无分润结果订单样本(前{len(no_share_samples)}条):')
                        for sample in no_share_samples:
                            task.add_log(sample, 'warning')
                            logger.warning(sample)
                task.progress = 90
                db.session.commit()
                _update_progress(task_id, 90, '生成导出文件...')

                # 7. 准备导出数据
                export_rows = []
                for (agent_no, month), share in sorted(monthly_shares.items(), key=lambda x: (x[0][1], x[0][0])):
                    node = agent_info_map.get(agent_no)
                    if node:
                        export_rows.append({
                            '代理商编号': agent_no,
                            '代理商名称': node.agent_name,
                            '分润金额': round(share, 2),
                            '月份': month,
                            '代理商电话': node.login_phone,
                            '代理商等级': node.rank,
                            '所属一级代理商名称': node.root_name,
                            '上级代理商': node.parent_name,
                            '上级关系': agent_ancestor_chain.get(agent_no, ''),
                        })

                if not export_rows:
                    task.add_log('未生成分润数据', 'warning')

                # 8. 生成 Excel
                task.add_log('生成Excel文件...')
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_filename = f'profit_share_{org_no}_{timestamp}.xlsx'
                output_path = os.path.join(output_dir, output_filename)

                _generate_excel(export_rows, output_path, order_detail_rows)

                task.add_log(f'导出完成，共 {len(export_rows)} 行分润记录，{len(order_detail_rows)} 行订单明细')
                task.status = 'completed'
                task.completed_at = datetime.utcnow()
                task.output_file = output_path
                task.total_rows = len(export_rows)
                task.success_count = len(export_rows)
                task.failure_count = 0
                task.progress = 100
                task.error_message = None
                db.session.commit()
                _update_progress(task_id, 100, '导出完成')

                if on_complete:
                    try:
                        on_complete(task_id, 'completed')
                    except Exception:
                        pass

            except Exception as e:
                logger.error(f"代理分润导出失败: {str(e)}", exc_info=True)
                task.status = 'failed'
                task.completed_at = datetime.utcnow()
                task.error_message = str(e)
                task.progress = 100
                task.add_log(f'导出失败: {str(e)}', 'error')
                db.session.commit()
                _update_progress(task_id, 100, f'导出失败: {str(e)}', 'error')
                if on_complete:
                    try:
                        on_complete(task_id, 'failed')
                    except Exception:
                        pass

    @staticmethod
    def get_task_status(task_id: str) -> Optional[dict]:
        task = QueryTask.query.filter_by(task_id=task_id).first()
        if not task:
            return None
        result = task.to_dict()
        progress_info = _get_progress(task_id)
        if progress_info.get('progress', 0) > result.get('progress', 0):
            result['progress'] = progress_info['progress']
        return result

    @staticmethod
    def cancel_task(task_id: str) -> bool:
        task = QueryTask.query.filter_by(task_id=task_id, type='export').first()
        if not task:
            return False
        if task.status in ('pending', 'running'):
            cancel_event = _task_cancel_events.get(task_id)
            if cancel_event:
                cancel_event.set()

            thread = _task_threads.get(task_id)
            if thread and thread.is_alive():
                try:
                    import ctypes
                    tid = ctypes.c_long(thread.ident)
                    ctypes.pythonapi.PyThreadState_SetAsyncExc(tid, ctypes.py_object(SystemExit))
                except Exception:
                    pass

            task.status = 'manual_cancelled'
            task.completed_at = datetime.utcnow()
            task.add_log('任务已被手动终止', 'warning')
            db.session.commit()

            _task_threads.pop(task_id, None)
            _task_cancel_events.pop(task_id, None)
            return True
        return False
