import logging
import os
import uuid
import threading
from datetime import datetime
from typing import Any, Dict, List, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
from app import db
from app.models.database import DatabaseConnection
from app.models.script import Script
from app.models.query_task import QueryTask
from app.utils.db_connector import DatabaseConnector
from app.utils.sql_validator import SQLValidator
from app.services.excel_service import ExcelService
from app.utils.helpers import beijing_isoformat
from app.utils.sql_template import render_sql_template

logger = logging.getLogger(__name__)

_task_progress = {}
_task_lock = threading.Lock()


def update_task_progress(task_id: str, progress: int, log_message: str = None, level: str = 'info'):
    with _task_lock:
        _task_progress[task_id] = {
            'progress': progress,
            'log_message': log_message,
            'level': level,
            'timestamp': beijing_isoformat(datetime.utcnow())
        }


def get_task_progress(task_id: str) -> dict:
    with _task_lock:
        return _task_progress.get(task_id, {'progress': 0})


class QueryService:

    @staticmethod
    def create_task(script_id: int, db_connection_ids: List[int], input_file: str,
                    merge_strategy: str = 'concat') -> QueryTask:
        task_id = str(uuid.uuid4())
        task = QueryTask(
            task_id=task_id,
            script_id=script_id,
            database_connection_id=db_connection_ids[0] if db_connection_ids else None,
            input_file=input_file,
            status='pending',
            merge_strategy=merge_strategy,
        )
        task.set_database_ids(db_connection_ids)
        db.session.add(task)
        db.session.commit()
        return task

    @staticmethod
    def execute_query_async(task_id: str, script_id: int, script_ids: List[int],
                            db_connection_ids: List[int], input_file: str, output_dir: str,
                            param_column: str = None, merge_strategy: str = 'concat',
                            new_sheet: bool = True, column_mapping: dict = None,
                            primary_key: str = ''):
        thread = threading.Thread(
            target=QueryService._execute_query_background,
            args=(task_id, script_id, script_ids, db_connection_ids, input_file, output_dir,
                  param_column, merge_strategy, new_sheet, column_mapping, primary_key),
            daemon=True
        )
        thread.start()

    @staticmethod
    def _execute_query_background(task_id: str, script_id: int, script_ids: List[int],
                                  db_connection_ids: List[int], input_file: str, output_dir: str,
                                  param_column: str = None, merge_strategy: str = 'concat',
                                  new_sheet: bool = True, column_mapping: dict = None,
                                  primary_key: str = ''):
        try:
            from app import create_app
            app = create_app()
        except Exception:
            from flask import current_app
            app = current_app._get_current_object()

        with app.app_context():
            task = QueryTask.query.filter_by(task_id=task_id).first()
            if not task:
                return

            try:
                task.status = 'running'
                task.started_at = datetime.utcnow()
                task.add_log('开始执行查询任务')
                task.progress = 5
                db.session.commit()
                update_task_progress(task_id, 5, '开始执行查询任务')

                all_results = {}
                total_scripts = len(script_ids)
                param_name = param_column

                excel_data = ExcelService.read_params(input_file, param_column=param_name)
                params_data = excel_data['data']
                column_names = excel_data['column_names']

                if not param_name and column_names:
                    param_name = column_names[0]

                if not params_data:
                    raise ValueError("Excel文件中没有数据")

                params_list = [{param_name: str(val)} for val in params_data]
                task.total_rows = len(params_list)
                task.add_log(f'从Excel读取到 {len(params_list)} 条参数，参数列: {param_name}')
                task.progress = 10
                db.session.commit()
                update_task_progress(task_id, 10, f'读取到{len(params_list)}条参数')

                for script_idx, sid in enumerate(script_ids):
                    script = Script.query.get(sid)
                    if not script:
                        task.add_log(f'脚本不存在: {sid}', 'warning')
                        continue

                    validator = SQLValidator()
                    validation = validator.validate(script.sql_text)
                    if not validation.is_valid:
                        task.add_log(f'SQL验证失败 [{script.name}]: {validation.message}', 'error')
                        continue

                    task.add_log(f'执行查询选项 [{script.name}]，模式: {script.query_mode}')

                    # 如果启用了SQL模板，先渲染模板生成实际SQL
                    sql_to_execute = script.sql_text
                    if script.is_template and script.sql_template:
                        try:
                            template_config = script.get_template_config()
                            sql_to_execute = render_sql_template(script.sql_template, template_config)
                            task.add_log(f'查询选项 [{script.name}] SQL模板渲染完成')
                        except Exception as e:
                            task.add_log(f'查询选项 [{script.name}] SQL模板渲染失败: {str(e)}', 'error')
                            continue
                    else:
                        sql_to_execute = script.sql_text

                    script_db_ids = script.get_database_ids()
                    if not script_db_ids:
                        script_db_ids = db_connection_ids

                    connectors = {}
                    for conn_id in script_db_ids:
                        conn_model = DatabaseConnection.query.get(conn_id)
                        if not conn_model:
                            continue
                        try:
                            connector = DatabaseConnector(conn_model.to_config_dict())
                            if connector.test_connection():
                                connectors[conn_id] = {
                                    'connector': connector,
                                    'name': conn_model.name,
                                    'model': conn_model
                                }
                                task.add_log(f'数据库连接成功: {conn_model.name}')
                            else:
                                task.add_log(f'数据库连接测试失败: {conn_model.name}', 'warning')
                        except Exception as e:
                            task.add_log(f'数据库连接失败 {conn_model.name}: {str(e)}', 'error')

                    if not connectors:
                        task.add_log(f'查询选项 [{script.name}] 没有可用的数据库连接', 'error')
                        continue

                    progress_base = 10 + int(70 * script_idx / total_scripts)
                    progress_range = int(70 / total_scripts)

                    with ThreadPoolExecutor(max_workers=min(len(connectors), 8)) as executor:
                        futures = {}
                        for conn_id, conn_info in connectors.items():
                            future = executor.submit(
                                QueryService._execute_on_database,
                                conn_id, conn_info['connector'], conn_info['name'],
                                sql_to_execute, params_list, script.query_mode,
                                script.timeout, script.batch_size
                            )
                            futures[future] = conn_id

                        completed = 0
                        for future in as_completed(futures):
                            conn_id = futures[future]
                            try:
                                result = future.result()
                                result_key = f"{sid}_{conn_id}"
                                all_results[result_key] = result
                                completed += 1
                                progress = progress_base + int(progress_range * completed / len(connectors))
                                task.progress = min(progress, 85)
                                total_result_rows = sum(
                                    len(item.get('result', []))
                                    for r in all_results.values()
                                    for item in r.get('results', []) if r.get('success')
                                )
                                task.add_log(f'[{script.name}] 数据库 {result["db_name"]} 查询完成: 成功{result["success_count"]}, 失败{result["failure_count"]}, 结果行{total_result_rows}')
                                db.session.commit()
                                update_task_progress(task_id, task.progress, f'数据库 {result["db_name"]} 查询完成')
                            except Exception as e:
                                completed += 1
                                conn_info = connectors[conn_id]
                                result_key = f"{sid}_{conn_id}"
                                all_results[result_key] = {
                                    'db_name': conn_info['name'],
                                    'success': False,
                                    'results': [],
                                    'errors': [{'error': str(e)}],
                                    'success_count': 0,
                                    'failure_count': len(params_list)
                                }
                                task.add_log(f'数据库查询失败: {str(e)}', 'error')
                                db.session.commit()

                    for conn_id, conn_info in connectors.items():
                        try:
                            conn_info['connector'].close()
                        except Exception:
                            pass

                task.progress = 85
                task.add_log('开始处理查询结果并写入Excel')
                db.session.commit()
                update_task_progress(task_id, 85, '处理查询结果')

                first_script = Script.query.get(script_ids[0]) if script_ids else None
                column_headers = []
                if first_script:
                    # 模板模式下从渲染后的SQL提取列名
                    if first_script.is_template and first_script.sql_template:
                        try:
                            template_config = first_script.get_template_config()
                            rendered_sql = render_sql_template(first_script.sql_template, template_config)
                            column_headers = SQLValidator().extract_column_names(rendered_sql)
                        except Exception:
                            column_headers = SQLValidator().extract_column_names(first_script.sql_text)
                    else:
                        column_headers = SQLValidator().extract_column_names(first_script.sql_text)

                processed_results = QueryService._process_all_results(
                    all_results, column_headers, param_name,
                    first_script.query_mode if first_script else 'batch'
                )

                total_success = sum(r.get('success_count', 0) for r in all_results.values())
                total_failure = sum(r.get('failure_count', 0) for r in all_results.values())

                # 查询成功但未查询到数据，不生成结果文件
                if len(processed_results) == 0:
                    task.status = 'completed'
                    task.completed_at = datetime.utcnow()
                    task.total_rows = 0
                    task.output_file = None
                    task.success_count = total_success
                    task.failure_count = total_failure
                    task.progress = 100
                    task.error_message = None
                    task.add_log('查询执行完成，但未查询到任何数据，未生成结果文件', 'warning')
                    db.session.commit()
                    update_task_progress(task_id, 100, '查询执行完成（无数据）')
                    return

                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_filename = f"result_{timestamp}.xlsx"
                output_path = os.path.join(output_dir, output_filename)

                if new_sheet:
                    QueryService._write_new_sheet(
                        output_path, input_file, processed_results,
                        column_headers, column_mapping or {},
                        first_script.result_sheet_name if first_script else '查询结果',
                        merge_strategy, all_results, {}
                    )
                else:
                    QueryService._write_existing_sheet_with_mapping(
                        output_path, input_file, processed_results,
                        param_name, column_mapping or {}, primary_key
                    )

                task.status = 'completed'
                task.completed_at = datetime.utcnow()
                task.output_file = output_path
                task.success_count = total_success
                task.failure_count = total_failure
                task.progress = 100
                task.add_log(f'查询执行完成，结果已写入: {output_filename}')
                db.session.commit()
                update_task_progress(task_id, 100, '查询执行完成')

            except Exception as e:
                logger.error(f"查询执行失败: {str(e)}", exc_info=True)
                task.status = 'failed'
                task.completed_at = datetime.utcnow()
                task.error_message = str(e)
                task.progress = 100
                task.add_log(f'查询执行失败: {str(e)}', 'error')
                db.session.commit()
                update_task_progress(task_id, 100, f'查询执行失败: {str(e)}', 'error')

    @staticmethod
    def _execute_on_database(conn_id: int, connector: DatabaseConnector, db_name: str,
                             sql: str, params_list: List[Dict], query_mode: str,
                             timeout: int, batch_size: int, max_rows: int = 0) -> Dict[str, Any]:
        try:
            if query_mode == 'in':
                result = connector.execute_in_query(sql, params_list, timeout=timeout, max_rows=max_rows)
            else:
                result = connector.execute_batch_queries(sql, params_list, timeout=timeout, batch_size=batch_size, max_rows=max_rows)

            return {
                'db_id': conn_id,
                'db_name': db_name,
                'success': True,
                'results': result.get('results', []),
                'errors': result.get('errors', []),
                'success_count': result.get('success', 0),
                'failure_count': result.get('failure', 0),
            }
        except Exception as e:
            return {
                'db_id': conn_id,
                'db_name': db_name,
                'success': False,
                'results': [],
                'errors': [{'error': str(e)}],
                'success_count': 0,
                'failure_count': len(params_list),
            }

    @staticmethod
    def _process_all_results(all_results: Dict[str, Any], column_headers: List[str],
                             param_name: str, query_mode: str) -> List[Dict[str, Any]]:
        processed = []

        for conn_id, db_result in all_results.items():
            if not db_result.get('success'):
                continue

            for item in db_result.get('results', []):
                if not item.get('success', False):
                    continue
                rows = item.get('result', [])
                params = item.get('params', {})

                for row in rows:
                    row_dict = {}
                    if column_headers and len(column_headers) == len(row):
                        for i, header in enumerate(column_headers):
                            row_dict[header] = row[i]
                    elif column_headers and len(row) > 0:
                        for i in range(min(len(column_headers), len(row))):
                            row_dict[column_headers[i]] = row[i]
                        for i in range(len(column_headers), len(row)):
                            row_dict[f'extra_col_{i}'] = row[i]
                    else:
                        for i, val in enumerate(row):
                            row_dict[f'col_{i}'] = val

                    row_dict['_db_name'] = db_result.get('db_name', '')
                    processed.append(row_dict)

        return processed

    @staticmethod
    def _write_new_sheet(output_path: str, input_file: str, results: List[Dict],
                         column_headers: List[str], column_mapping: dict,
                         sheet_name: str, merge_strategy: str,
                         all_results: Dict, connectors: Dict):
        import shutil
        from app.utils.excel_writer import ExcelWriter

        shutil.copy2(input_file, output_path)
        writer = ExcelWriter(output_path)

        try:
            if not writer.load_existing():
                writer.create_new()

            if merge_strategy == 'separate':
                for conn_id, db_result in all_results.items():
                    if not db_result.get('success') or not db_result.get('results'):
                        continue
                    db_name = db_result.get('db_name', f'db_{conn_id}')
                    db_sheet_name = f"{sheet_name}_{db_name}"
                    db_results = [r for r in results if r.get('_db_name') == db_name]
                    if db_results:
                        QueryService._write_results_to_sheet(writer, db_sheet_name, db_results, column_headers, column_mapping)
            else:
                display_results = []
                for r in results:
                    r_copy = {k: v for k, v in r.items() if k != '_db_name'}
                    display_results.append(r_copy)
                QueryService._write_results_to_sheet(writer, sheet_name, display_results, column_headers, column_mapping)

            writer.save()
        finally:
            writer.close()

    @staticmethod
    def _write_results_to_sheet(writer, sheet_name: str, results: List[Dict],
                                column_headers: List[str], column_mapping: dict):
        if not results:
            return

        writer.create_sheet(sheet_name)

        headers = []
        sql_keys = []
        if column_headers:
            for h in column_headers:
                mapped = column_mapping.get(h, h) if column_mapping else h
                headers.append(mapped)
                sql_keys.append(h)
        else:
            headers = list(results[0].keys())
            sql_keys = headers[:]

        writer.write_headers(headers)

        write_chunk_size = 5000
        for chunk_start in range(0, len(results), write_chunk_size):
            chunk = results[chunk_start:chunk_start + write_chunk_size]
            data = []
            for row in chunk:
                row_data = [row.get(k, '') for k in sql_keys]
                data.append(row_data)
            start_row = chunk_start + 2  # +1 for 0-based index, +1 for header row
            writer.write_data(data, start_row=start_row)

    @staticmethod
    def _write_existing_sheet(output_path: str, input_file: str, results: List[Dict],
                              param_column_name: str):
        import shutil
        from app.utils.excel_writer import ExcelWriter

        shutil.copy2(input_file, output_path)
        writer = ExcelWriter(output_path)

        try:
            if writer.load_existing():
                writer.get_active_sheet()
                display_results = [{k: v for k, v in r.items() if k != '_db_name'} for r in results]
                writer.write_data_by_column_name(display_results, param_column_name)
                writer.save()
        finally:
            writer.close()

    @staticmethod
    def _get_synonym_groups():
        """从系统配置获取同义词组，如果没有配置则使用默认值"""
        default_groups = [
            ['编号', '号', '代码', '编码', 'ID', 'id', 'No', 'no'],
            ['名称', '名', '名字', 'NAME', 'name'],
            ['金额', '数额', '额度', '数目'],
            ['日期', '时间', 'Date', 'date', 'Time', 'time'],
            ['商户', '商户号', '商户编号'],
            ['我方', '我司', '系统', '本方'],
            ['对方', '他方', '渠道', '通道'],
            ['注册', '入驻', '登记'],
            ['手机号', '手机号码', '电话', '手机', '电话号码'],
            ['名称', '姓名', '名字']
        ]
        try:
            from app.models.system_config import SystemConfig
            config = SystemConfig.query.filter_by(
                config_key=SystemConfig.COLUMN_SYNONYM_GROUPS
            ).first()
            if config and config.config_value:
                import json
                groups = json.loads(config.config_value)
                if isinstance(groups, list) and groups:
                    return groups
        except Exception:
            pass
        return default_groups

    @staticmethod
    def _fuzzy_match_columns(sql_fields: List[str], excel_headers: List[str]) -> dict:
        """
        模糊匹配SQL字段与Excel列名。
        匹配优先级：
        1. 完全匹配（忽略大小写和空格）
        2. 去除常见后缀/前缀后的匹配（如"编号"↔"号"，"名称"↔"名"）
        3. 包含匹配（SQL字段包含在Excel列名中，或反之）
        4. 核心关键词匹配（提取字段中的核心词，在Excel列名中查找包含该核心词的列）
        """
        mapping = {}
        used_excel_cols = set()

        # 从系统配置获取同义词映射
        synonym_groups = QueryService._get_synonym_groups()

        def get_synonyms(word):
            """获取一个词的所有同义词"""
            for group in synonym_groups:
                if word in group:
                    return group
            return [word]

        def extract_core_parts(field):
            """提取字段名的核心部分，去掉常见前缀/后缀"""
            import re
            # 去掉表名前缀 (如 t1. / table.)
            core = re.sub(r'^\w+\.', '', field)
            # 去掉引号
            core = core.replace('`', '').replace('"', '').replace('[', '').replace(']', '')
            return core.strip()

        def normalize(s):
            """标准化：去空格、转小写"""
            return s.replace(' ', '').replace('_', '').replace('-', '').lower()

        # 第一轮：完全匹配（忽略大小写和空格）
        for sql_field in sql_fields:
            core = extract_core_parts(sql_field)
            norm_core = normalize(core)
            for eh in excel_headers:
                if eh in used_excel_cols:
                    continue
                if normalize(eh) == norm_core:
                    mapping[sql_field] = eh
                    used_excel_cols.add(eh)
                    break

        # 第二轮：同义词+包含匹配
        remaining_sql = [f for f in sql_fields if f not in mapping]
        if remaining_sql:
            for sql_field in remaining_sql:
                core = extract_core_parts(sql_field)
                norm_core = normalize(core)
                # 将SQL字段名拆分为关键词
                sql_keywords = [core]
                for group in synonym_groups:
                    for word in group:
                        if word in core:
                            # 用同义词组中的所有词替换，生成可能的变体
                            for syn in group:
                                if syn != word:
                                    variant = core.replace(word, syn)
                                    sql_keywords.append(variant)
                            break

                best_match = None
                best_score = 0

                for eh in excel_headers:
                    if eh in used_excel_cols:
                        continue
                    norm_eh = normalize(eh)

                    # 检查SQL字段的各变体是否与Excel列匹配
                    for keyword in sql_keywords:
                        norm_kw = normalize(keyword)
                        if norm_kw == norm_eh:
                            best_match = eh
                            best_score = 100
                            break
                        # 包含匹配：SQL字段包含在Excel列名中
                        if norm_kw and norm_eh and (norm_kw in norm_eh or norm_eh in norm_kw):
                            score = min(len(norm_kw), len(norm_eh)) / max(len(norm_kw), len(norm_eh), 1) * 80
                            if score > best_score:
                                best_score = score
                                best_match = eh

                    if best_score == 100:
                        break

                if best_match and best_score >= 50:
                    mapping[sql_field] = best_match
                    used_excel_cols.add(best_match)

        # 第三轮：核心关键词匹配（从SQL字段中提取核心词，在Excel列名中查找）
        remaining_sql = [f for f in sql_fields if f not in mapping]
        if remaining_sql:
            # 从同义词组中提取所有词作为潜在的前缀/后缀修饰词
            all_modifier_words = set()
            for group in synonym_groups:
                for word in group:
                    all_modifier_words.add(word)

            for sql_field in remaining_sql:
                core = extract_core_parts(sql_field)
                norm_core = normalize(core)

                # 提取核心关键词（去掉修饰词后）
                core_words = [core]
                # 尝试去掉前缀修饰词（从同义词组中提取）
                for modifier in all_modifier_words:
                    if core.startswith(modifier):
                        remaining = core[len(modifier):]
                        if remaining:
                            core_words.append(remaining)
                # 尝试去掉后缀修饰词（从同义词组中提取）
                for modifier in all_modifier_words:
                    if core.endswith(modifier):
                        remaining = core[:-len(modifier)]
                        if remaining:
                            core_words.append(remaining)

                best_match = None
                best_score = 0

                for eh in excel_headers:
                    if eh in used_excel_cols:
                        continue
                    norm_eh = normalize(eh)

                    for cw in core_words:
                        norm_cw = normalize(cw)
                        if not norm_cw:
                            continue
                        # 检查核心词是否被包含
                        if norm_cw in norm_eh or norm_eh in norm_cw:
                            score = min(len(norm_cw), len(norm_eh)) / max(len(norm_cw), len(norm_eh), 1) * 60
                            if score > best_score:
                                best_score = score
                                best_match = eh

                if best_match and best_score >= 40:
                    mapping[sql_field] = best_match
                    used_excel_cols.add(best_match)

        logger.info(f"模糊列匹配结果: {mapping}")
        return mapping

    @staticmethod
    def _write_existing_sheet_with_mapping(output_path: str, input_file: str, results: List[Dict],
                                           param_column_name: str, column_mapping: dict,
                                           primary_key: str = ''):
        import shutil
        import openpyxl

        shutil.copy2(input_file, output_path)

        try:
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active

            excel_headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                excel_headers.append(str(cell.value) if cell.value else '')

            if not results:
                wb.save(output_path)
                return

            # 自动模糊匹配：对column_mapping中未映射的字段进行自动匹配
            if results:
                sql_fields = list(results[0].keys())
                unmapped = [f for f in sql_fields if f not in column_mapping or not column_mapping[f]]
                if unmapped:
                    fuzzy_mapping = QueryService._fuzzy_match_columns(unmapped, excel_headers)
                    for k, v in fuzzy_mapping.items():
                        if k not in column_mapping or not column_mapping[k]:
                            column_mapping[k] = v
                    if fuzzy_mapping:
                        logger.info(f"自动模糊匹配补充列映射: {fuzzy_mapping}")

            # 主键列也需要模糊匹配
            pk_sql_col = primary_key
            pk_excel_col = column_mapping.get(primary_key, primary_key)

            if primary_key and (not pk_excel_col or pk_excel_col not in excel_headers):
                # 尝试模糊匹配主键列
                pk_fuzzy = QueryService._fuzzy_match_columns([primary_key], excel_headers)
                if primary_key in pk_fuzzy:
                    pk_excel_col = pk_fuzzy[primary_key]
                    column_mapping[primary_key] = pk_excel_col
                    logger.info(f"主键列模糊匹配: {primary_key} -> {pk_excel_col}")

            reverse_mapping = {}
            for sql_col, excel_col in column_mapping.items():
                if excel_col and excel_col != '隐藏':
                    reverse_mapping[excel_col] = sql_col

            logger.info(f"列映射写入: column_mapping={column_mapping}")
            logger.info(f"列映射写入: primary_key={primary_key}, pk_sql_col={pk_sql_col}, pk_excel_col={pk_excel_col}")
            logger.info(f"列映射写入: Excel headers={excel_headers}")
            logger.info(f"列映射写入: results count={len(results)}, first_result_keys={list(results[0].keys()) if results else 'empty'}")

            if primary_key and pk_excel_col in excel_headers:
                pk_col_idx = excel_headers.index(pk_excel_col) + 1

                pk_row_map = {}
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    cell_val = row[pk_col_idx - 1].value
                    pk_val = str(cell_val).strip() if cell_val is not None else ''
                    if pk_val:
                        pk_row_map[pk_val] = row_idx

                logger.info(f"列映射写入(主键模式): 主键Excel列={pk_excel_col}, 主键SQL列={pk_sql_col}, "
                           f"Excel中找到{len(pk_row_map)}个主键值, 查询结果{len(results)}条")
                if pk_row_map:
                    logger.info(f"列映射写入: Excel主键样本={list(pk_row_map.keys())[:5]}")

                matched = 0
                for result in results:
                    pk_cell_val = result.get(pk_sql_col)
                    pk_value = str(pk_cell_val).strip() if pk_cell_val is not None else ''
                    if pk_value and pk_value in pk_row_map:
                        row_num = pk_row_map[pk_value]
                        for sql_col, excel_col in column_mapping.items():
                            if not excel_col or excel_col == '隐藏':
                                continue
                            if sql_col not in result:
                                continue
                            col_idx = None
                            for i, h in enumerate(excel_headers):
                                if h == excel_col:
                                    col_idx = i + 1
                                    break
                            if col_idx:
                                ws.cell(row=row_num, column=col_idx, value=result[sql_col])
                        matched += 1

                logger.info(f"列映射写入完成: 匹配{matched}行")
            else:
                logger.info(f"列映射写入(追加模式): 无主键或主键列[{pk_excel_col}]不在Excel headers中")

                for result in results:
                    row_data = []
                    for h in excel_headers:
                        sql_col = reverse_mapping.get(h)
                        if sql_col and sql_col in result:
                            row_data.append(result[sql_col])
                        else:
                            row_data.append(result.get(h, ''))
                    ws.append(row_data)

            wb.save(output_path)
        except Exception as e:
            logger.error(f"写入Excel失败: {str(e)}")
            raise

    @staticmethod
    def get_task_status(task_id: str) -> Optional[Dict[str, Any]]:
        task = QueryTask.query.filter_by(task_id=task_id).first()
        if not task:
            return None
        result = task.to_dict()
        progress_info = get_task_progress(task_id)
        if progress_info.get('progress', 0) > result.get('progress', 0):
            result['progress'] = progress_info['progress']
        script = Script.query.get(task.script_id) if task.script_id else None
        result['script_name'] = script.name if script else '未知'
        result['script_tag'] = script.tag if script else ''
        db_names = []
        for db_id in task.get_database_ids():
            conn = DatabaseConnection.query.get(db_id)
            if conn:
                db_names.append(conn.name)
        result['database_names'] = db_names
        return result

    @staticmethod
    def cancel_task(task_id: str) -> bool:
        task = QueryTask.query.filter_by(task_id=task_id).first()
        if not task:
            return False
        if task.status in ('pending', 'running'):
            task.status = 'cancelled'
            task.completed_at = datetime.utcnow()
            task.add_log('任务已取消', 'warning')
            db.session.commit()
            update_task_progress(task_id, 100, '任务已取消', 'warning')
            return True
        return False

    @staticmethod
    def get_dashboard_stats(user_id=None, is_admin=False) -> Dict[str, Any]:
        total_databases = DatabaseConnection.query.filter_by(is_active=True).count()
        total_scripts = Script.query.filter_by(is_active=True, type='query').count()
        base_q = QueryTask.query
        if not is_admin and user_id:
            base_q = base_q.filter_by(created_by=user_id)
        total_tasks = base_q.count()
        recent_tasks = base_q.order_by(QueryTask.created_at.desc()).limit(5).all()
        success_tasks = base_q.filter_by(status='completed').count()
        failed_tasks = base_q.filter_by(status='failed').count()

        recent_list = []
        for task in recent_tasks:
            task_dict = task.to_dict()
            script = Script.query.get(task.script_id) if task.script_id else None
            task_dict['script_name'] = script.name if script else '未知'
            task_dict['script_tag'] = script.tag if script else ''
            db_names = []
            for db_id in task.get_database_ids():
                conn = DatabaseConnection.query.get(db_id)
                if conn:
                    db_names.append(conn.name)
            task_dict['database_names'] = db_names
            recent_list.append(task_dict)

        return {
            'total_databases': total_databases,
            'total_scripts': total_scripts,
            'total_tasks': total_tasks,
            'success_tasks': success_tasks,
            'failed_tasks': failed_tasks,
            'recent_tasks': recent_list,
        }
