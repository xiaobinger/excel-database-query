"""代付提现服务

集成原"代付提现"桌面工具的 4 个渠道逻辑：
- 合利宝(helipay)：代付
- 电银(dianyin)：代付 / 查询
- 乐商通PLUS(lepass)：代付(实时) / 查询
- 快乐刷(kls)：实时代付 / 跑批代付(创建/查询/发起提现) / 查询

Excel 列映射（固定列索引，从 0 开始）：
  row[2]=姓名 accountName, row[3]=流水号 businessNo, row[4]=手机号,
  row[5]=身份证号 idCardNo, row[6]=银行卡号 bankCardNo,
  row[8]=金额(分) amount, row[12]=子代理编号 objectDstId,
  row[13]=提现applyId, row[14]=代付响应/reqId, row[15]=提现响应,
  row[16]=查询响应, row[17]=提现响应
"""
import json
import random
import string
import time
import hashlib
import requests
import os
import xlrd
from openpyxl import load_workbook, Workbook


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------

def _cell(row, idx):
    """安全读取行数据（不足列数返回 None）"""
    if idx < len(row):
        v = row[idx]
        return v
    return None


def _to_str(v):
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _to_int(v):
    try:
        return int(float(v))
    except Exception:
        return 0


def format_amount_to_yuan(amount_fen):
    return "{:.2f}".format(amount_fen / 100)


def generate_unique_id(merchant_id, length=32):
    timestamp = int(time.time())
    characters = string.ascii_letters + string.digits
    random_part = ''.join(random.choices(characters, k=length - len(str(timestamp)) - len(str(merchant_id))))
    return f"{merchant_id}{timestamp}{random_part}"


def load_channels():
    """加载所有渠道元数据及其配置状态，供 AI 工具调用"""
    from app.models.pay_config import PayConfig
    CHANNEL_META = [
        {'channel': 'helipay', 'name': '合利宝', 'interface_types': ['代付'], 'real_time': False, 'execute_types': []},
        {'channel': 'dianyin', 'name': '电银', 'interface_types': ['代付', '查询'], 'real_time': False, 'execute_types': []},
        {'channel': 'lepass', 'name': '乐商通PLUS', 'interface_types': ['代付', '查询'], 'real_time': True, 'execute_types': []},
        {'channel': 'kls', 'name': '快乐刷', 'interface_types': ['代付', '查询'], 'real_time': True,
         'execute_types': ['创建代付', '查询代付', '发起提现']},
    ]
    rows = []
    for meta in CHANNEL_META:
        cfg_db = PayConfig.query.filter_by(channel=meta['channel']).first()
        item = dict(meta)
        item['has_config'] = cfg_db is not None and bool(cfg_db.get_test_config() or cfg_db.get_pro_config())
        rows.append(item)
    return rows


def get_config(channel):
    """获取指定渠道的配置（优先 test 环境），供 AI 工具调用"""
    from app.models.pay_config import PayConfig
    cfg_db = PayConfig.query.filter_by(channel=channel).first()
    if not cfg_db:
        return None
    return cfg_db.get_test_config() or cfg_db.get_pro_config()


def get_sheets_info(file_path):
    """获取 Excel 文件所有工作表信息（名称 + 数据行数），返回 list[dict]"""
    ext = os.path.splitext(file_path)[1].lower()
    sheets = []
    if ext == '.xlsx':
        wb = load_workbook(file_path, data_only=True)
        for ws in wb.worksheets:
            # 跳过表头行，统计数据行数
            data_rows = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(c is None or str(c).strip() == '' for c in row):
                    continue
                data_rows += 1
            sheets.append({'name': ws.title, 'row_count': data_rows})
    elif ext == '.xls':
        book = xlrd.open_workbook(file_path)
        for sheet in book.sheets():
            data_rows = max(sheet.nrows - 1, 0)  # 减去表头行
            sheets.append({'name': sheet.name, 'row_count': data_rows})
    return sheets


def _load_rows(file_path, sheet_index=0):
    """读取 Excel 数据行（跳过表头），返回 list[list]，支持指定工作表索引"""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.xlsx':
        wb = load_workbook(file_path, data_only=True)
        ws = wb.worksheets[sheet_index] if sheet_index < len(wb.worksheets) else wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(c is None or str(c).strip() == '' for c in row):
                continue
            rows.append(list(row))
        return rows
    elif ext == '.xls':
        book = xlrd.open_workbook(file_path)
        sheet = book.sheets()[sheet_index] if sheet_index < len(book.sheets()) else book.sheets()[0]
        rows = []
        for r in range(1, sheet.nrows):
            row_data = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
            if all(v is None or str(v).strip() == '' for v in row_data):
                continue
            rows.append(row_data)
        return rows
    return []


def _write_result(file_path, rows, result_rows, sheet_name='代付结果'):
    """生成结果 Excel：原始列 + 结果列，返回新文件路径"""
    import os
    out_path = file_path + '_result.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # 表头：原始表头无法直接获取（已跳过），用列序号占位 + 结果列
    ws.append([f'列{i+1}' for i in range(len(rows[0])) if rows[0]] + ['子代理编号', 'applyId', '代付响应', '提现响应', '查询响应', '提现查询响应'])
    for r in result_rows:
        ws.append(r)
    wb.save(out_path)
    return out_path


def _post_json(url, payload, timeout=30, log=None):
    """POST JSON 请求，打印详细请求和响应日志"""
    if log:
        try:
            payload_str = json.dumps(payload, ensure_ascii=False)
        except (TypeError, ValueError):
            payload_str = str(payload)
        log(f'[HTTP请求] POST {url}')
        log(f'[HTTP请求] Payload: {payload_str}')
    try:
        resp = requests.post(url, json=payload, timeout=timeout)
        if log:
            log(f'[HTTP响应] 状态码: {resp.status_code}')
            log(f'[HTTP响应] Body: {resp.text[:2000]}')
        return resp
    except Exception as e:
        if log:
            log(f'[HTTP异常] {type(e).__name__}: {e}')
        raise


# ---------------------------------------------------------------------------
# 合利宝
# ---------------------------------------------------------------------------

def helipay_pay(cfg, rows, params, log):
    success, fail, success_amount = 0, 0, 0
    result_rows = []
    for row in rows:
        business_no = _to_str(_cell(row, 3))
        account_name = _to_str(_cell(row, 2))
        bank_card_no = _to_str(_cell(row, 6))
        amount_fen = _to_int(_cell(row, 8))
        amount_yuan = format_amount_to_yuan(amount_fen)

        payload = {
            'agentNo': cfg.get('agentNo'),
            'data': {
                'agentNo': cfg.get('agentNo'),
                'userId': cfg.get('userId'),
                'withdrawAmount': amount_yuan,
                'cardNo': bank_card_no,
                'payerName': account_name,
                'bankCode': params.get('bank_code', 'CCB'),
                'orderNum': business_no,
                'onlineBankType': params.get('online_bank_type', 'B2C'),
            },
            'userId': cfg.get('userId'),
            'url': cfg.get('helipayUrl', '') + '/mpos-trx/rest/withdraw/agentApply.action',
        }
        try:
            resp = _post_json(cfg.get('baseUrl', '') + '/test/heliPay/sendRequest', payload, log=log)
            text = resp.text
            result_json = json.loads(text)
            if result_json.get('retCode') == '0000':
                success += 1
                success_amount += amount_fen
                resp_str = '成功'
            else:
                fail += 1
                resp_str = text[:500]
            log(f'合利宝 流水{business_no} 金额{amount_yuan}元 => {resp_str}')
        except Exception as e:
            fail += 1
            resp_str = f'异常: {e}'
            log(f'合利宝 流水{business_no} 异常: {e}')
        result_rows.append(row + ['' , '', resp_str, '', '', ''])
    message = f'合利宝代付成功：{success}笔, 失败：{fail}笔, 成功金额：{format_amount_to_yuan(success_amount)}元'
    return message, result_rows


# ---------------------------------------------------------------------------
# 电银
# ---------------------------------------------------------------------------

def _dianyin_sign(json_obj, key):
    data = json.loads(json_obj) if isinstance(json_obj, str) else dict(json_obj)
    sorted_data = dict(sorted(data.items()))
    result = "&".join([f"{k}={v}" for k, v in sorted_data.items()])
    result = result + "key=" + key
    return hashlib.sha256(result.encode('utf-8')).hexdigest()


def _dianyin_request(cfg, url, signature, param_json, log=None):
    """电银请求：POST + signature header，打印详细请求和响应日志"""
    headers = {'Content-Type': 'application/json', 'signature': signature}
    if log:
        log(f'[电银HTTP请求] POST {url}')
        log(f'[电银HTTP请求] Headers: signature={signature}')
        log(f'[电银HTTP请求] Body: {param_json if isinstance(param_json, str) else json.dumps(param_json, ensure_ascii=False)}')
    try:
        resp = requests.post(url, data=param_json, headers=headers, timeout=30)
        if log:
            log(f'[电银HTTP响应] 状态码: {resp.status_code}')
            log(f'[电银HTTP响应] Body: {resp.text[:2000]}')
        return resp
    except Exception as e:
        if log:
            log(f'[电银HTTP异常] {type(e).__name__}: {e}')
        raise


def dianyin_withdraw(cfg, order_no, amount, account_json, log=None):
    """账户报备 + 发起提现"""
    account_report_obj = json.loads(account_json)
    account_report_obj["drawType"] = 'FROZEN'
    account_report_obj["partnerIden"] = 'dyin'
    account_report_obj["channelNo"] = cfg.get('channelNo', '81429675')
    account_report_obj["partnerNo"] = cfg.get('partnerNo', '88805')
    account_report_obj = json.dumps(account_report_obj)

    if log:
        log(f'[电银] 步骤1: 账户报备 流水{order_no}')
    account_report_url = cfg.get('accountReportUrl', 'http://39.101.182.160:8070/api/with-draw/account-contract')
    account_report_sign = _dianyin_sign(account_report_obj, cfg.get('signKey', ''))
    account_report_result = _dianyin_request(cfg, account_report_url, account_report_sign, account_report_obj, log)

    if account_report_result.status_code == 200:
        account_report_json = account_report_result.json()
        if account_report_json.get("code") == 200:
            account_id = account_report_json["data"]["accId"]
            if log:
                log(f'[电银] 账户报备成功 accId={account_id}')
            withdraw_obj = {
                "accId": account_id,
                "amount": amount,
                "asSupportInvoice": "N",
                "channelNo": cfg.get('channelNo', '81429675'),
                "drawType": "FROZEN",
                "orderNo": order_no,
                "partnerIden": "dyin",
                "partnerNo": cfg.get('partnerNo', '88805'),
            }
            if log:
                log(f'[电银] 步骤2: 发起提现 流水{order_no} 金额{amount}')
            withdraw_url = cfg.get('withdrawUrl', 'http://39.101.182.160:8070/api/with-draw/with-draw')
            withdraw_result = _dianyin_request(cfg, withdraw_url, _dianyin_sign(json.dumps(withdraw_obj), cfg.get('signKey', '')), json.dumps(withdraw_obj), log)
            if withdraw_result.json().get("code") == 200 and withdraw_result.json().get("data", {}).get("orderStatus") == 'ONTHEWAY':
                return True
            if log:
                log(f'[电银] 提现失败: {withdraw_result.text[:500]}')
            return False
        if log:
            log(f'[电银] 账户报备失败: {account_report_json}')
        return False
    if log:
        log(f'[电银] 账户报备HTTP错误: {account_report_result.status_code}')
    return False


def dianyin_query(cfg, order_no, log=None):
    query_obj = {
        "channelNo": cfg.get('channelNo', '81429675'),
        "orderNo": order_no,
        "partnerIden": "dyin",
        "partnerNo": cfg.get('partnerNo', '88805'),
    }
    query_url = cfg.get('queryUrl', 'http://39.101.182.160:8070/api/with-draw/order-query')
    return _dianyin_request(cfg, query_url, _dianyin_sign(json.dumps(query_obj), cfg.get('signKey', '')), json.dumps(query_obj), log)


def dianyin_pay(cfg, rows, params, log):
    success, fail, success_amount = 0, 0, 0
    result_rows = []
    for row in rows:
        amount = _to_str(_cell(row, 8))
        order_no = _to_str(_cell(row, 3))
        id_name = _to_str(_cell(row, 2))
        bank_phone = _to_str(_cell(row, 4))
        bank_account = _to_str(_cell(row, 6))
        id_card_no = _to_str(_cell(row, 5))
        account_report_obj = {
            "bankAccount": bank_account,
            "bankCityCode": cfg.get('bankCityCode', '371700'),
            "bankDistrictCode": cfg.get('bankDistrictCode', '371702'),
            "bankPhone": bank_phone,
            "bankProvinceCode": cfg.get('bankProvinceCode', '370000'),
            "branchName": cfg.get('branchName', '中国工商银行'),
            "idCardNo": id_card_no,
            "idName": id_name,
        }
        try:
            flag = dianyin_withdraw(cfg, order_no, amount, json.dumps(account_report_obj, ensure_ascii=False), log)
            if flag:
                success += 1
                success_amount += _to_int(amount)
                resp_str = '成功'
            else:
                fail += 1
                resp_str = '报备/提现失败'
            log(f'电银 流水{order_no} 金额{format_amount_to_yuan(_to_int(amount))}元 => {resp_str}')
        except Exception as e:
            fail += 1
            resp_str = f'异常: {e}'
            log(f'电银 流水{order_no} 异常: {e}')
        result_rows.append(row + ['', '', resp_str, '', '', ''])
    message = f'电银代付成功：{success}笔, 金额：{format_amount_to_yuan(success_amount)}元, 失败：{fail}笔'
    return message, result_rows


def dianyin_query_batch(cfg, rows, params, log):
    success, fail, paying = 0, 0, 0
    fail_flows = ''
    result_rows = []
    for row in rows:
        order_no = _to_str(_cell(row, 3))
        try:
            result = dianyin_query(cfg, order_no, log)
            data = result.json()
            if data.get("code") == 200:
                status = data.get("data", {}).get("orderStatus")
                if status == 'SUCCESS':
                    success += 1
                    resp_str = '成功'
                elif status == 'FAIL':
                    fail += 1
                    fail_flows += f'{order_no}({data.get("data", {}).get("remitMsg")}), '
                    resp_str = f'失败: {data.get("data", {}).get("remitMsg")}'
                else:
                    paying += 1
                    resp_str = f'打款中: {status}'
            else:
                fail += 1
                fail_flows += f'{order_no}({data.get("msg")}), '
                resp_str = f'查询失败: {data.get("msg")}'
            log(f'电银查询 流水{order_no} => {resp_str}')
        except Exception as e:
            fail += 1
            resp_str = f'异常: {e}'
            log(f'电银查询 流水{order_no} 异常: {e}')
        result_rows.append(row + ['', '', '', '', resp_str, ''])
    message = f'电银代付成功：{success}笔, 失败：{fail}笔, 失败流水：{fail_flows}, 打款中：{paying}笔'
    return message, result_rows


# ---------------------------------------------------------------------------
# 乐商通PLUS / 快乐刷（共用 external-api 接口）
# ---------------------------------------------------------------------------

def _get_sub_agent_id(cfg, bank_card_no, channel_code, log=None):
    url = cfg.get('getSubAgentIdUrl', '')
    payload = {"bankCardNo": bank_card_no, "channelCode": channel_code}
    if log:
        log(f'[获取子代理] POST {url} 银行卡={bank_card_no} 渠道={channel_code}')
    resp = requests.post(url, json=payload, timeout=30)
    if log:
        log(f'[获取子代理] 状态码: {resp.status_code} 响应: {resp.text[:500]}')
    return resp.text.strip()


def _lsp_transfer(cfg, business_no, object_dst_id, amount, transfer_mode, log=None):
    payload = {
        "source": cfg.get("source"),
        "objectSrcId": cfg.get("objectSrcId"),
        "objectDstId": object_dst_id,
        "amount": amount,
        "transferMode": transfer_mode,
        "key": cfg.get("key"),
        "businessNo": business_no,
        "sign": cfg.get("sign"),
    }
    return _post_json(cfg.get("baseUrl", '') + '/external-api/behalfPay/transfer', payload, log=log).text


def _lsp_create_pay(cfg, business_no, object_dst_id, amount, busi_type, log=None):
    payload = {
        "source": cfg.get("source"),
        "agentId": cfg.get("objectSrcId"),
        "subAgentId": object_dst_id,
        "amount": amount,
        "busiType": busi_type,
        "key": cfg.get("key"),
        "reqId": business_no,
        "sign": cfg.get("sign"),
    }
    return _post_json(cfg.get("baseUrl", '') + '/external-api/behalfPay/bPay', payload, log=log).text


def _lsp_query_pay(cfg, req_id, busi_type, log=None):
    payload = {
        "source": cfg.get("source"),
        "agentId": cfg.get("objectSrcId"),
        "busiType": busi_type,
        "key": cfg.get("key"),
        "reqId": req_id,
        "sign": cfg.get("sign"),
    }
    return _post_json(cfg.get("baseUrl", '') + '/external-api/behalfPay/queryV2', payload, log=log).text


def _lsp_withdraw(cfg, business_no, object_dst_id, amount, log=None):
    payload = {
        "agentId": cfg.get("objectSrcId"),
        "applyAgentId": object_dst_id,
        "accountType": "1",
        "sign": cfg.get("sign"),
        "invoiceType": "0",
        "source": cfg.get("source"),
        "businessNo": business_no,
        "applyAmount": amount,
        "key": cfg.get("key"),
    }
    return _post_json(cfg.get("baseUrl", '') + '/external-api/merchantinfo/v3/submitWithdraw', payload, log=log).text


def _lsp_query_order(cfg, business_no, log=None):
    payload = {"source": cfg.get("source"), "key": cfg.get("key"), "businessNo": business_no, "sign": cfg.get("sign")}
    return _post_json(cfg.get("baseUrl", '') + '/external-api/behalfPay/transferOrderInfo', payload, log=log).text


def _lsp_query_withdraw(cfg, object_dst_id, apply_id, log=None):
    payload = {
        "source": cfg.get("source"),
        "key": cfg.get("key"),
        "applyId": apply_id,
        "withdrawAgentNo": object_dst_id,
        "agentId": cfg.get("objectSrcId"),
        "sign": cfg.get("sign"),
    }
    return _post_json(cfg.get("baseUrl", '') + '/external-api/merchantinfo/v3/queryWithdrawState', payload, log=log).text


def _check_without_agent(cfg, id_card_no):
    without = cfg.get("withoutAgents", "") or ""
    return id_card_no in without


def lepass_realtime_pay(cfg, rows, params, log):
    """乐商通PLUS / 快乐刷 实时代付（transferMode 区分）"""
    transfer_mode = params.get('transfer_mode') or '7'
    channel_code = params.get('channel_code') or 'lepass'
    success, fail, success_amount = 0, 0, 0
    result_rows = []
    for row in rows:
        id_card_no = _to_str(_cell(row, 5))
        if _check_without_agent(cfg, id_card_no):
            log(f'跳过 身份证{id_card_no} 历史有出款异常')
            result_rows.append(row + ['', '', '历史有出款异常,跳过', '', '', ''])
            continue
        business_no = generate_unique_id(_to_str(_cell(row, 3)))
        bank_card_no = _to_str(_cell(row, 6))
        object_dst_id = _get_sub_agent_id(cfg, bank_card_no, channel_code, log)
        if not bank_card_no or not object_dst_id:
            result_rows.append(row + ['', '', '需要创建子代理', '', '', ''])
            log(f'流水{_to_str(_cell(row, 3))} 需要创建子代理')
            continue
        amount = _to_int(_cell(row, 8))
        try:
            pay_result = json.loads(_lsp_transfer(cfg, business_no, object_dst_id, amount, transfer_mode, log))
            if pay_result.get("error_code") == 0 and pay_result.get("data", {}).get("state") == 2:
                order_info = _lsp_query_order(cfg, business_no, log)
                withdraw_biz_no = business_no + str(amount)
                withdraw_result = json.loads(_lsp_withdraw(cfg, withdraw_biz_no, object_dst_id, amount, log))
                if withdraw_result.get("error_code") == 0 and withdraw_result.get("error_msg") == "成功":
                    success += 1
                    success_amount += amount
                    apply_id = withdraw_result.get("data", {}).get("applyId")
                    withdraw_query = _lsp_query_withdraw(cfg, object_dst_id, apply_id, log)
                    result_rows.append(row + [object_dst_id, apply_id, order_info, withdraw_result.get("error_msg"), '', withdraw_query])
                    log(f'{channel_code} 流水{business_no} 代付+提现成功 applyId={apply_id}')
                else:
                    fail += 1
                    result_rows.append(row + [object_dst_id, '', order_info, withdraw_result.get("error_msg"), '', ''])
                    log(f'{channel_code} 流水{business_no} 提现失败: {withdraw_result.get("error_msg")}')
            else:
                fail += 1
                result_rows.append(row + [object_dst_id, '', pay_result.get("error_msg"), '', '', ''])
                log(f'{channel_code} 流水{business_no} 代付失败: {pay_result.get("error_msg")}')
        except Exception as e:
            fail += 1
            result_rows.append(row + [object_dst_id, '', f'异常: {e}', '', '', ''])
            log(f'{channel_code} 流水{business_no} 异常: {e}')
    message = f'代付提现成功：{success}笔, 失败：{fail}笔, 成功金额：{format_amount_to_yuan(success_amount)}元'
    return message, result_rows


def kls_create_pay(cfg, rows, params, log):
    """快乐刷 跑批-创建代付"""
    busi_type = params.get('busi_type') or '144'
    channel_code = params.get('channel_code') or 'kls'
    success, fail, success_amount = 0, 0, 0
    result_rows = []
    for row in rows:
        business_no = generate_unique_id(_to_str(_cell(row, 3)))
        bank_card_no = _to_str(_cell(row, 6))
        object_dst_id = _get_sub_agent_id(cfg, bank_card_no, channel_code, log)
        if not bank_card_no or not object_dst_id:
            result_rows.append(row + ['', '', '需要创建子代理', '', '', ''])
            continue
        amount = _to_int(_cell(row, 8))
        try:
            pay_result = json.loads(_lsp_create_pay(cfg, business_no, object_dst_id, amount, busi_type, log))
            if pay_result.get("error_code") == 0:
                success += 1
                success_amount += amount
                req_id = str(pay_result.get("data"))
                result_rows.append(row + [object_dst_id, '', req_id, pay_result.get("error_msg"), '', ''])
                log(f'快乐刷创建代付 流水{business_no} 成功 reqId={req_id}')
            else:
                fail += 1
                result_rows.append(row + [object_dst_id, '', '', pay_result.get("error_msg"), '', ''])
                log(f'快乐刷创建代付 流水{business_no} 失败: {pay_result.get("error_msg")}')
        except Exception as e:
            fail += 1
            result_rows.append(row + [object_dst_id, '', '', f'异常: {e}', '', ''])
            log(f'快乐刷创建代付 流水{business_no} 异常: {e}')
    message = f'快乐刷创建代付成功：{success}笔, 失败：{fail}笔, 成功金额：{format_amount_to_yuan(success_amount)}元'
    return message, result_rows


def kls_query_pay(cfg, rows, params, log):
    """快乐刷 跑批-查询代付"""
    busi_type = params.get('busi_type') or '144'
    success, fail, paying = 0, 0, 0
    result_rows = []
    for row in rows:
        req_id = _to_str(_cell(row, 14))
        if not req_id:
            result_rows.append(row + ['', '', '', '', '创建代付未成功', ''])
            continue
        try:
            pay_result = json.loads(_lsp_query_pay(cfg, req_id, busi_type, log))
            if pay_result.get("error_code") == 0:
                state = int(pay_result.get("data", {}).get("state"))
                if state == 3:
                    success += 1
                    resp_str = '成功'
                elif state in (2, 4):
                    fail += 1
                    resp_str = '失败'
                else:
                    paying += 1
                    resp_str = f'代付中({state})'
            else:
                fail += 1
                resp_str = pay_result.get("error_msg")
            result_rows.append(row + ['', '', '', '', resp_str, pay_result.get("error_msg")])
            log(f'快乐刷查询代付 reqId={req_id} => {resp_str}')
        except Exception as e:
            fail += 1
            result_rows.append(row + ['', '', '', '', f'异常: {e}', ''])
            log(f'快乐刷查询代付 reqId={req_id} 异常: {e}')
    message = f'快乐刷查询代付 成功：{success}笔, 失败：{fail}笔, 代付中：{paying}笔'
    return message, result_rows


def kls_apply_withdraw(cfg, rows, params, log):
    """快乐刷 跑批-发起提现"""
    channel_code = params.get('channel_code') or 'kls'
    success, fail, success_amount = 0, 0, 0
    result_rows = []
    for row in rows:
        business_no = generate_unique_id(_to_str(_cell(row, 3)))
        amount = _to_int(_cell(row, 8))
        object_dst_id = _to_str(_cell(row, 12))
        try:
            withdraw_result = json.loads(_lsp_withdraw(cfg, business_no, object_dst_id, amount, log))
            if withdraw_result.get("error_code") == 0 and withdraw_result.get("error_msg") == "成功":
                success += 1
                success_amount += amount
                apply_id = withdraw_result.get("data", {}).get("applyId")
                withdraw_query = _lsp_query_withdraw(cfg, object_dst_id, apply_id, log)
                result_rows.append(row + [object_dst_id, apply_id, '', withdraw_result.get("error_msg"), '', withdraw_query])
                log(f'快乐刷发起提现 流水{business_no} 成功 applyId={apply_id}')
            else:
                fail += 1
                result_rows.append(row + [object_dst_id, '', '', withdraw_result.get("error_msg"), '', ''])
                log(f'快乐刷发起提现 流水{business_no} 失败: {withdraw_result.get("error_msg")}')
        except Exception as e:
            fail += 1
            result_rows.append(row + [object_dst_id, '', '', f'异常: {e}', '', ''])
            log(f'快乐刷发起提现 流水{business_no} 异常: {e}')
    message = f'快乐刷发起提现成功：{success}笔, 失败：{fail}笔, 成功金额：{format_amount_to_yuan(success_amount)}元'
    return message, result_rows


def lsp_withdraw_query(cfg, rows, params, log):
    """乐商通PLUS / 快乐刷 提现状态查询（按 applyId）"""
    success, fail, paying = 0, 0, 0
    result_rows = []
    for row in rows:
        object_dst_id = _to_str(_cell(row, 12))
        apply_id = _to_str(_cell(row, 13))
        if not apply_id:
            result_rows.append(row + ['', '', '', '', '', '无applyId'])
            continue
        try:
            result = json.loads(_lsp_query_withdraw(cfg, object_dst_id, apply_id, log))
            if result.get("error_code") == 0:
                f_state = result.get("data", {}).get("F_state")
                if f_state in (-1, "6", 6):
                    paying += 1
                    resp_str = '打款中'
                elif f_state == "9":
                    success += 1
                    resp_str = '成功'
                else:
                    fail += 1
                    resp_str = f'失败({f_state})'
            else:
                fail += 1
                resp_str = result.get("error_msg")
            result_rows.append(row + ['', '', '', '', '', resp_str])
            log(f'提现查询 applyId={apply_id} => {resp_str}')
        except Exception as e:
            fail += 1
            result_rows.append(row + ['', '', '', '', '', f'异常: {e}'])
            log(f'提现查询 applyId={apply_id} 异常: {e}')
    message = f'打款成功：{success}笔, 失败：{fail}笔, 打款中：{paying}笔'
    return message, result_rows


# ---------------------------------------------------------------------------
# 统一入口
# ---------------------------------------------------------------------------

def execute_pay(channel, cfg, rows, params, log):
    """根据渠道 + 接口类型 + 跑批步骤分发执行，返回 (message, result_rows)"""
    interface_type = params.get('interface_type', '代付')
    real_time = params.get('real_time', '是')
    execute_type = params.get('execute_type', '创建代付')

    if channel == 'helipay':
        return helipay_pay(cfg, rows, params, log)
    elif channel == 'dianyin':
        if interface_type == '查询':
            return dianyin_query_batch(cfg, rows, params, log)
        return dianyin_pay(cfg, rows, params, log)
    elif channel in ('lepass', 'kls'):
        if interface_type == '查询':
            return lsp_withdraw_query(cfg, rows, params, log)
        # 代付
        if channel == 'lepass':
            return lepass_realtime_pay(cfg, rows, params, log)
        # 快乐刷
        if real_time == '是':
            return lepass_realtime_pay(cfg, rows, params, log)
        # 跑批
        if execute_type == '创建代付':
            return kls_create_pay(cfg, rows, params, log)
        elif execute_type == '查询代付':
            return kls_query_pay(cfg, rows, params, log)
        else:
            return kls_apply_withdraw(cfg, rows, params, log)
    else:
        raise ValueError(f'未知渠道: {channel}')


def execute_pay_withdraw(channel, interface_type, environment, file_path, sheet_index=0,
                         real_time='是', execute_type='创建代付', output_dir=None, on_complete=None):
    """供工单AI调用的异步执行入口（在后台线程中运行，需自行创建 app context）

    参数:
        channel: 渠道标识
        interface_type: 接口类型（代付/查询）
        environment: 环境（test/pro）
        file_path: Excel 文件路径
        sheet_index: 工作表索引
        real_time: 实时代付（是/否，仅快乐刷）
        execute_type: 跑批步骤（仅快乐刷跑批）
        output_dir: 结果文件输出目录
        on_complete: 回调函数 fn(status, message, logs, result_url=None)
            status: 'completed' 或 'failed'
            message: 执行结果摘要
            logs: 日志列表
            result_url: 结果文件下载URL（可选）
    """
    import logging
    _logger = logging.getLogger(__name__)
    logs = []

    def log(msg):
        logs.append(msg)
        _logger.info(f'[代付-工单] {msg}')

    try:
        from flask import current_app
        from app.models.pay_config import PayConfig

        cfg_db = PayConfig.query.filter_by(channel=channel).first()
        if not cfg_db:
            raise Exception(f'渠道 {channel} 未配置')

        cfg = cfg_db.get_pro_config() if environment == 'pro' else cfg_db.get_test_config()
        if not cfg:
            raise Exception(f'渠道 {channel} 的 {environment} 环境未配置')

        params = {
            'interface_type': interface_type,
            'real_time': real_time,
            'execute_type': execute_type,
            'bank_code': cfg_db.bank_code or 'CCB',
            'online_bank_type': cfg_db.online_bank_type or 'B2C',
            'transfer_mode': cfg_db.transfer_mode or ('6' if channel == 'kls' else '7'),
            'busi_type': cfg_db.busi_type or '144',
            'channel_code': cfg_db.channel_code or ('kls' if channel == 'kls' else 'lepass'),
        }

        rows = _load_rows(file_path, sheet_index=sheet_index)
        if not rows:
            raise Exception(f'Excel 工作表索引 {sheet_index} 无数据行')

        sheets = get_sheets_info(file_path)
        sheet_name = sheets[sheet_index]['name'] if sheet_index < len(sheets) else str(sheet_index)
        env_label = '生产' if environment == 'pro' else '测试'
        log(f'开始执行代付提现：渠道={channel} 环境={env_label} 接口={interface_type} 工作表={sheet_name} 数据行数={len(rows)}')

        message, result_rows = execute_pay(channel, cfg, rows, params, log)
        result_path = _write_result(file_path, rows, result_rows)
        result_name = os.path.basename(result_path)

        # 移动结果文件到输出目录
        result_url = None
        if output_dir and os.path.exists(result_path):
            os.makedirs(output_dir, exist_ok=True)
            dest_path = os.path.join(output_dir, result_name)
            import shutil
            shutil.move(result_path, dest_path)
            result_url = f'/api/download/ticket/{result_name}'

        log(f'执行完成：{message}')
        if on_complete:
            on_complete('completed', message, logs, result_url)
    except Exception as e:
        _logger.exception('代付提现执行失败')
        log(f'执行失败: {e}')
        if on_complete:
            on_complete('failed', str(e), logs, None)
