import logging
import os
import re
import uuid
import zipfile
import threading
from datetime import datetime
from typing import Any, Dict, List, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
from app import db
from app.models.database import DatabaseConnection
from app.models.script import Script
from app.models.query_task import QueryTask
from app.utils.db_connector import DatabaseConnector
from app.utils.sql_validator import SQLValidator
from app.utils.helpers import beijing_isoformat
from app.utils.sql_template import render_sql_template

logger = logging.getLogger(__name__)

_task_progress = {}
_task_lock = threading.Lock()


def update_export_progress(task_id: str, progress: int, log_message: str = None, level: str = 'info'):
    with _task_lock:
        _task_progress[task_id] = {
            'progress': progress,
            'log_message': log_message,
            'level': level,
            'timestamp': beijing_isoformat(datetime.utcnow())
        }


def get_export_progress(task_id: str) -> dict:
    with _task_lock:
        return _task_progress.get(task_id, {'progress': 0})


class ExportService:

    @staticmethod
    def create_export_task(script_ids: List[int], params_values: Dict[str, Any],
                           output_format: str = 'sheets', created_by: int = None) -> QueryTask:
        task_id = str(uuid.uuid4())
        task = QueryTask(
            task_id=task_id,
            script_id=script_ids[0] if script_ids else None,
            status='pending',
            type='export',
            output_format=output_format,
        )
        task.set_script_ids_json(script_ids)
        task.set_params_values(params_values)

        all_db_ids = []
        for sid in script_ids:
            script = Script.query.get(sid)
            if script:
                db_ids = script.get_database_ids()
                all_db_ids.extend(db_ids)
        all_db_ids = list(set(all_db_ids))
        task.set_database_ids(all_db_ids)
        if all_db_ids:
            task.database_connection_id = all_db_ids[0]

        if created_by:
            task.created_by = created_by

        db.session.add(task)
        db.session.commit()
        return task

    @staticmethod
    def execute_export_async(task_id: str, script_ids: List[int],
                             params_values: Dict[str, Any], output_dir: str,
                             output_format: str = 'sheets', on_complete=None):
        thread = threading.Thread(
            target=ExportService._execute_export_background,
            args=(task_id, script_ids, params_values, output_dir, output_format, on_complete),
            daemon=True
        )
        thread.start()

    @staticmethod
    def _execute_export_background(task_id: str, script_ids: List[int],
                                   params_values: Dict[str, Any], output_dir: str,
                                   output_format: str = 'sheets', on_complete=None):
        try:
            from app import create_app
            app = create_app()
        except Exception:
            from flask import current_app
            app = current_app._get_current_object()

        with app.app_context():
            task = QueryTask.query.filter_by(task_id=task_id).first()
            if not task:
                return

            try:
                task.status = 'running'
                task.started_at = datetime.utcnow()
                task.add_log('开始执行导出任务')
                task.progress = 5
                db.session.commit()
                update_export_progress(task_id, 5, '开始执行导出任务')

                all_results = {}
                total_scripts = len(script_ids)

                for script_idx, sid in enumerate(script_ids):
                    script = Script.query.get(sid)
                    if not script:
                        task.add_log(f'导出选项不存在: {sid}', 'warning')
                        continue

                    sql_text = script.sql_text

                    # 如果启用了SQL模板，先渲染模板生成实际SQL
                    if script.is_template and script.sql_template:
                        try:
                            template_config = script.get_template_config()
                            rendered_sql = render_sql_template(script.sql_template, template_config)
                            sql_text = rendered_sql
                            task.add_log(f'导出选项 [{script.name}] SQL模板渲染完成')
                        except Exception as e:
                            task.add_log(f'导出选项 [{script.name}] SQL模板渲染失败: {str(e)}', 'error')
                            continue

                    script_params = dict(params_values) if params_values else {}
                    params_config = script.get_params_config()
                    multi_params = {p['name'] for p in params_config if p.get('multi') and p.get('type') in ('text', 'number')}
                    number_params = {p['name'] for p in params_config if p.get('type') == 'number'}
                    neq_params = {p['name']: p for p in params_config if p.get('enum_enabled') and p.get('enum_mode') == 'neq' and p.get('neq_value')}
                    allow_all_params = {p['name'] for p in params_config if p.get('enum_enabled') and p.get('allow_all')}

                    # 1. 先处理"全部"选项：智能移除WHERE条件
                    for pname in allow_all_params:
                        if pname not in script_params:
                            # 匹配模式：column = {{param}} 或 column != {{param}}
                            # 场景1: WHERE column = {{param}} → 整个WHERE去掉
                            # 场景2: WHERE ... AND column = {{param}} → 去掉 AND ...
                            # 场景3: WHERE column = {{param}} AND ... → 去掉 column = {{param}} AND
                            col_pat = rf'`?{re.escape(pname)}`?'
                            cond_pat = rf'({col_pat}\s*[=!]=\s*\{{\{{\s*{re.escape(pname)}\s*\}}\}}\s*)'
                            
                            # 场景A: 只有这一个WHERE条件 → 整条WHERE去掉
                            sql_text = re.sub(
                                rf'\bWHERE\s+{cond_pat}(?=\s*(?:;|$|GROUP|ORDER|LIMIT|UNION))',
                                '',
                                sql_text,
                                flags=re.IGNORECASE
                            )
                            # 场景B: WHERE cond AND ... → 删cond AND，保留WHERE给后续条件
                            sql_text = re.sub(
                                rf'(?<=\bWHERE\b)\s+{cond_pat}\s+AND\s+',
                                ' ',
                                sql_text,
                                flags=re.IGNORECASE
                            )
                            # 场景C: ... AND cond → 删AND cond（覆盖中间和末尾位置）
                            sql_text = re.sub(
                                rf'\s+AND\s+{cond_pat}',
                                '',
                                sql_text,
                                flags=re.IGNORECASE
                            )

                    # 2. 处理非即不等于参数（选是/否时替换=为!=）
                    for pname, pconf in neq_params.items():
                        if pname in script_params:
                            is_checked = script_params[pname]
                            neq_val = pconf.get('neq_value', '')
                            if is_checked:
                                script_params[pname] = neq_val
                            else:
                                # 勾选否：将 '= {{param}}' 整体替换为 '!= neq_val'
                                neq_val_fmt = f"!= {neq_val}" if pname in number_params else f"!= '{neq_val}'"
                                sql_text = re.sub(
                                    rf'=\s*\{{\{{\s*{re.escape(pname)}\s*\}}\}}\s*',
                                    neq_val_fmt,
                                    sql_text
                                )
                                del script_params[pname]

                    if script_params:
                        for param_name in multi_params:
                            if param_name not in script_params:
                                continue
                            val = str(script_params[param_name])
                            parts = [v.strip() for v in val.split(',') if v.strip()]
                            if not parts:
                                continue
                            if len(parts) == 1:
                                fmt = f"'{parts[0]}'" if param_name not in number_params else parts[0]
                                sql_text = sql_text.replace(f'{{{{{param_name}}}}}', fmt)
                                if f':{param_name}' in sql_text:
                                    script_params[param_name] = parts[0]
                            else:
                                quoted = ','.join(f"'{p}'" for p in parts) if param_name not in number_params else ','.join(parts)
                                sql_text = sql_text.replace(f'{{{{{param_name}}}}}', quoted)
                                if f':{param_name}' in sql_text:
                                    bind_keys = [f':{param_name}_{i}' for i in range(len(parts))]
                                    sql_text = re.sub(
                                        rf':{param_name}\b',
                                        ', '.join(bind_keys),
                                        sql_text
                                    )
                                    for i, part in enumerate(parts):
                                        script_params[f'{param_name}_{i}'] = part
                                    del script_params[param_name]

                        placeholders = re.findall(r'\{\{(\w+)\}\}', sql_text)
                        for ph in placeholders:
                            if ph in script_params:
                                val = script_params[ph]
                                val_str = str(val)
                                if ph in number_params:
                                    sql_text = sql_text.replace(f'{{{{{ph}}}}}', val_str)
                                elif val_str and not val_str.startswith("'"):
                                    sql_text = sql_text.replace(f'{{{{{ph}}}}}', f"'{val_str}'")
                                else:
                                    sql_text = sql_text.replace(f'{{{{{ph}}}}}', val_str)
                        placeholders_remaining = re.findall(r'\{\{(\w+)\}\}', sql_text)
                        for ph in placeholders_remaining:
                            sql_text = sql_text.replace(f'{{{{{ph}}}}}', '')

                    task.add_log(f'执行导出选项 [{script.name}]')

                    script_db_ids = script.get_database_ids()
                    if not script_db_ids:
                        task.add_log(f'导出选项 [{script.name}] 没有关联数据库连接', 'error')
                        continue

                    connectors = {}
                    for conn_id in script_db_ids:
                        conn_model = DatabaseConnection.query.get(conn_id)
                        if not conn_model:
                            continue
                        try:
                            connector = DatabaseConnector(conn_model.to_config_dict())
                            if connector.test_connection():
                                connectors[conn_id] = {
                                    'connector': connector,
                                    'name': conn_model.name,
                                    'model': conn_model
                                }
                                task.add_log(f'数据库连接成功: {conn_model.name}')
                            else:
                                task.add_log(f'数据库连接测试失败: {conn_model.name}', 'warning')
                        except Exception as e:
                            task.add_log(f'数据库连接失败 {conn_model.name}: {str(e)}', 'error')

                    if not connectors:
                        task.add_log(f'导出选项 [{script.name}] 没有可用的数据库连接', 'error')
                        continue

                    progress_base = 10 + int(60 * script_idx / total_scripts)
                    progress_range = int(60 / total_scripts)

                    for conn_id, conn_info in connectors.items():
                        try:
                            import re as _re
                            bind_params = _re.findall(r':(\w+)', sql_text)
                            if bind_params and script_params:
                                clean_params = {}
                                for k, v in script_params.items():
                                    if f':{k}' in sql_text and k in bind_params:
                                        clean_params[k] = v
                                if clean_params:
                                    result_rows = conn_info['connector'].execute_query(
                                        sql_text, clean_params, timeout=script.timeout, chunk_size=5000
                                    )
                                else:
                                    result_rows = conn_info['connector'].execute_query(
                                        sql_text, {}, timeout=script.timeout, chunk_size=5000
                                    )
                            else:
                                result_rows = conn_info['connector'].execute_query(
                                    sql_text, {}, timeout=script.timeout, chunk_size=5000
                                )

                            validator = SQLValidator()
                            # 模板模式下从渲染后的SQL提取列名
                            if script.is_template and script.sql_template:
                                try:
                                    template_config = script.get_template_config()
                                    rendered_for_cols = render_sql_template(script.sql_template, template_config)
                                    column_headers = validator.extract_column_names(rendered_for_cols)
                                except Exception:
                                    column_headers = validator.extract_column_names(sql_text)
                            else:
                                column_headers = validator.extract_column_names(sql_text)

                            rows_data = []
                            for row in result_rows:
                                row_dict = {}
                                if column_headers and len(column_headers) == len(row):
                                    for i, header in enumerate(column_headers):
                                        row_dict[header] = row[i]
                                elif column_headers and len(row) > 0:
                                    for i in range(min(len(column_headers), len(row))):
                                        row_dict[column_headers[i]] = row[i]
                                    for i in range(len(column_headers), len(row)):
                                        row_dict[f'extra_col_{i}'] = row[i]
                                else:
                                    for i, val in enumerate(row):
                                        row_dict[f'col_{i}'] = val
                                rows_data.append(row_dict)

                            result_key = f"{sid}_{conn_id}"
                            all_results[result_key] = {
                                'script_id': sid,
                                'script_name': script.name,
                                'db_name': conn_info['name'],
                                'columns': column_headers,
                                'rows': rows_data,
                                'success': True,
                            }
                            task.add_log(
                                f'[{script.name}] 数据库 {conn_info["name"]} 查询完成: {len(rows_data)} 行'
                            )
                        except Exception as e:
                            result_key = f"{sid}_{conn_id}"
                            all_results[result_key] = {
                                'script_id': sid,
                                'script_name': script.name,
                                'db_name': conn_info.get('name', ''),
                                'columns': [],
                                'rows': [],
                                'success': False,
                                'error': str(e),
                            }
                            task.add_log(f'[{script.name}] 数据库查询失败: {str(e)}', 'error')

                    for conn_id, conn_info in connectors.items():
                        try:
                            conn_info['connector'].close()
                        except Exception:
                            pass

                    progress = progress_base + progress_range
                    task.progress = min(progress, 75)
                    db.session.commit()
                    update_export_progress(task_id, task.progress, f'导出选项 [{script.name}] 完成')

                task.progress = 80
                db.session.commit()
                update_export_progress(task_id, 80, '检查查询结果')

                total_rows = sum(len(r.get('rows', [])) for r in all_results.values() if r.get('success'))
                total_success = sum(1 for r in all_results.values() if r.get('success'))
                total_failure = sum(1 for r in all_results.values() if not r.get('success'))

                if total_success == 0:
                    task.status = 'failed'
                    task.completed_at = datetime.utcnow()
                    task.total_rows = 0
                    task.success_count = 0
                    task.failure_count = total_failure
                    task.progress = 100
                    if total_failure > 0:
                        errors = [r.get('error', '') for r in all_results.values() if not r.get('success') and r.get('error')]
                        task.error_message = '所有查询均失败' + (f': {errors[0]}' if errors else '')
                    else:
                        task.error_message = '没有可用的查询结果'
                    task.add_log(f'导出失败: {task.error_message}', 'error')
                    db.session.commit()
                    update_export_progress(task_id, 100, f'导出失败: {task.error_message}', 'error')
                    if on_complete:
                        try:
                            on_complete(task_id, 'failed')
                        except Exception:
                            pass
                    return

                task.add_log('开始生成导出文件')
                db.session.commit()
                update_export_progress(task_id, 85, '生成导出文件')

                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

                try:
                    if output_format == 'zip':
                        output_path = ExportService._write_zip_output(
                            output_dir, timestamp, all_results, script_ids
                        )
                    else:
                        output_path = ExportService._write_sheets_output(
                            output_dir, timestamp, all_results, script_ids
                        )
                except Exception as file_err:
                    task.status = 'failed'
                    task.completed_at = datetime.utcnow()
                    task.error_message = f'生成文件失败: {str(file_err)}'
                    task.progress = 100
                    task.add_log(f'生成文件失败: {str(file_err)}', 'error')
                    db.session.commit()
                    update_export_progress(task_id, 100, f'生成文件失败: {str(file_err)}', 'error')
                    if on_complete:
                        try:
                            on_complete(task_id, 'failed')
                        except Exception:
                            pass
                    return

                task.status = 'completed'
                task.completed_at = datetime.utcnow()
                task.output_file = output_path
                task.total_rows = total_rows
                task.success_count = total_success
                task.failure_count = total_failure
                task.progress = 100
                task.add_log(f'导出完成，结果已写入: {os.path.basename(output_path)}')
                db.session.commit()
                update_export_progress(task_id, 100, '导出完成')
                if on_complete:
                    try:
                        on_complete(task_id, 'completed')
                    except Exception:
                        pass

            except Exception as e:
                logger.error(f"导出执行失败: {str(e)}", exc_info=True)
                task.status = 'failed'
                task.completed_at = datetime.utcnow()
                task.error_message = str(e)
                task.progress = 100
                task.add_log(f'导出执行失败: {str(e)}', 'error')
                db.session.commit()
                update_export_progress(task_id, 100, f'导出执行失败: {str(e)}', 'error')
                if on_complete:
                    try:
                        on_complete(task_id, 'failed')
                    except Exception:
                        pass

    @staticmethod
    def _write_sheets_output(output_dir: str, timestamp: str,
                             all_results: Dict, script_ids: List[int]) -> str:
        import openpyxl

        output_filename = f"export_{timestamp}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        wb = openpyxl.Workbook()
        first_sheet = True

        for sid in script_ids:
            script_results = {k: v for k, v in all_results.items() if v.get('script_id') == sid}
            if not script_results:
                continue

            script = Script.query.get(sid)
            sheet_name = script.name if script else f'导出_{sid}'
            sheet_name = sheet_name[:31]

            merged_rows = []
            merged_columns = []

            for result_key, result in script_results.items():
                if not result.get('success') or not result.get('rows'):
                    continue

                columns = result.get('columns', [])
                rows = result.get('rows', [])

                if not merged_columns and columns:
                    merged_columns = columns

                if columns and columns != merged_columns and merged_columns:
                    for row_data in rows:
                        mapped = {}
                        for i, col in enumerate(columns):
                            target_col = merged_columns[i] if i < len(merged_columns) else col
                            mapped[target_col] = row_data.get(col)
                        merged_rows.append(mapped)
                else:
                    merged_rows.extend(rows)

            if not merged_rows:
                continue

            if first_sheet:
                ws = wb.active
                ws.title = sheet_name
                first_sheet = False
            else:
                ws = wb.create_sheet(title=sheet_name)

            if merged_columns:
                for col_idx, col_name in enumerate(merged_columns, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)
            elif merged_rows:
                merged_columns = list(merged_rows[0].keys())
                for col_idx, col_name in enumerate(merged_columns, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)

            for row_idx, row_data in enumerate(merged_rows, 2):
                if merged_columns:
                    for col_idx, col_name in enumerate(merged_columns, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name))
                else:
                    for col_idx, val in enumerate(row_data.values(), 1):
                        ws.cell(row=row_idx, column=col_idx, value=val)

        if first_sheet:
            ws = wb.active
            ws.cell(row=1, column=1, value='无数据')

        wb.save(output_path)
        return output_path

    @staticmethod
    def _write_zip_output(output_dir: str, timestamp: str,
                          all_results: Dict, script_ids: List[int]) -> str:
        import openpyxl
        import tempfile

        zip_filename = f"export_{timestamp}.zip"
        zip_path = os.path.join(output_dir, zip_filename)

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for sid in script_ids:
                script_results = {k: v for k, v in all_results.items() if v.get('script_id') == sid}
                if not script_results:
                    continue

                script = Script.query.get(sid)
                base_name = script.name if script else f'export_{sid}'
                file_name = f"{base_name}.xlsx".replace(' ', '_')

                merged_rows = []
                merged_columns = []

                for result_key, result in script_results.items():
                    if not result.get('success') or not result.get('rows'):
                        continue

                    columns = result.get('columns', [])
                    rows = result.get('rows', [])

                    if not merged_columns and columns:
                        merged_columns = columns

                    if columns and columns != merged_columns and merged_columns:
                        for row_data in rows:
                            mapped = {}
                            for i, col in enumerate(columns):
                                target_col = merged_columns[i] if i < len(merged_columns) else col
                                mapped[target_col] = row_data.get(col)
                            merged_rows.append(mapped)
                    else:
                        merged_rows.extend(rows)

                if not merged_rows:
                    continue

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = base_name[:31]

                if merged_columns:
                    for col_idx, col_name in enumerate(merged_columns, 1):
                        ws.cell(row=1, column=col_idx, value=col_name)
                elif merged_rows:
                    merged_columns = list(merged_rows[0].keys())
                    for col_idx, col_name in enumerate(merged_columns, 1):
                        ws.cell(row=1, column=col_idx, value=col_name)

                for row_idx, row_data in enumerate(merged_rows, 2):
                    if merged_columns:
                        for col_idx, col_name in enumerate(merged_columns, 1):
                            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name))
                    else:
                        for col_idx, val in enumerate(row_data.values(), 1):
                            ws.cell(row=row_idx, column=col_idx, value=val)

                tmp_path = os.path.join(tempfile.gettempdir(), f"{uuid.uuid4().hex}.xlsx")
                wb.save(tmp_path)
                zf.write(tmp_path, file_name)
                os.remove(tmp_path)

        return zip_path

    @staticmethod
    def get_task_status(task_id: str) -> Optional[Dict[str, Any]]:
        task = QueryTask.query.filter_by(task_id=task_id).first()
        if not task:
            return None
        result = task.to_dict()
        progress_info = get_export_progress(task_id)
        if progress_info.get('progress', 0) > result.get('progress', 0):
            result['progress'] = progress_info['progress']

        script_ids = task.get_script_ids_json()
        script_names = []
        script_tags = []
        for sid in script_ids:
            script = Script.query.get(sid)
            if script:
                script_names.append(script.name)
                if script.tag:
                    script_tags.append(script.tag)

        result['script_names'] = script_names
        result['script_tags'] = script_tags

        db_names = []
        for db_id in task.get_database_ids():
            conn = DatabaseConnection.query.get(db_id)
            if conn:
                db_names.append(conn.name)
        result['database_names'] = db_names
        return result
